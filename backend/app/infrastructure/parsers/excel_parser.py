"""Parser de importação via Excel (RF010) — alternativa quando não há XML.

Campos mínimos (RF010): Nota Fiscal, Data Emissão, Transportadora, Cidade
Origem, Cidade Destino, Peso, Valor Mercadoria, Valor Frete.

Data Emissão é obrigatória (mesmo papel do ``dhEmi`` do XML): é a fonte da
competência (mês/ano) usada em todo indicador e filtro por período — Base de
CT-e por competência, Dashboard, DLG etc. Sem ela, o sistema não tem como
saber a que mês o frete pertence (planilhas antigas confundiam isso com a
data da importação, gerando competência incorreta).

Campos opcionais (compatibilidade retroativa — planilhas sem estas colunas
continuam funcionando normalmente):
  - Data Embarque / Data Saída / Data Entrega: datas operacionais do
    transporte (não são a data de emissão do CT-e).
  - CT-e: número do Conhecimento de Transporte.
  - Destinatário / CNPJ Destinatário: identifica o cliente final da carga
    (RN-67/RN-68 do DLG) — alimenta a dimensão CLIENTE_FINAL do diagnóstico
    (Diagnóstico Logístico e Recomendações). Sem essas colunas, os CT-es da
    planilha caem em "Cliente não identificado" (mesmo comportamento já
    usado para XML sem destinatário reconhecido — nenhum CT-e é descartado).
  - Componentes do frete: Frete Peso, Frete Valor, Pedágio, GRIS, Seguro,
    Ademe, Despacho, Outros (alimentam a composição usada na análise de
    divergência por componente, igual ao caminho de XML).
  - Tomador / CNPJ Tomador / Filial: CNPJ da matriz/filial responsável pelo
    frete — alimenta a dimensão FILIAL do DLG (mesmo papel do tomador do
    XML). Aceita só os dígitos ou um texto com o CNPJ embutido (ex.:
    "03822909000113 - Razão Social"), já que só os dígitos são extraídos
    (``_so_digitos``). Sem essa coluna, os CT-es caem em "Filial não
    identificada".
"""
from datetime import date, datetime
from typing import List, Optional

from openpyxl import load_workbook

from app.core.logging_config import get_logger

logger = get_logger(__name__)

# Mapeia variações de cabeçalho (normalizadas) -> campo canônico
COLUNAS = {
    "nota fiscal": "nota_fiscal",
    "nf": "nota_fiscal",
    "cte": "cte",
    "ct-e": "cte",
    "ct e": "cte",
    "numero cte": "cte",
    "numero do cte": "cte",
    "n cte": "cte",
    "conhecimento": "cte",
    "conhecimento de transporte": "cte",
    "data emissao": "data_emissao",
    "data emissão": "data_emissao",
    "data de emissao": "data_emissao",
    "data de emissão": "data_emissao",
    "emissao": "data_emissao",
    "emissão": "data_emissao",
    "data embarque": "data_embarque",
    "data saida": "data_saida",
    "data saída": "data_saida",
    "data entrega": "data_entrega",
    "destinatario": "destinatario_nome",
    "destinatário": "destinatario_nome",
    "cliente": "destinatario_nome",
    "cliente destinatario": "destinatario_nome",
    "cliente destinatário": "destinatario_nome",
    "destinatario cnpj": "destinatario_cnpj",
    "destinatário cnpj": "destinatario_cnpj",
    "cnpj destinatario": "destinatario_cnpj",
    "cnpj destinatário": "destinatario_cnpj",
    "cnpj cliente": "destinatario_cnpj",
    "tomador": "tomador_cnpj",
    "cnpj tomador": "tomador_cnpj",
    "tomador cnpj": "tomador_cnpj",
    "cnpj do tomador": "tomador_cnpj",
    "filial": "tomador_cnpj",
    "transportadora": "transportadora",
    "cidade origem": "cidade_origem",
    "cidade destino": "cidade_destino",
    "peso": "peso",
    "valor mercadoria": "valor_mercadoria",
    "valor frete": "valor_frete",
    "uf origem": "uf_origem",
    "uf destino": "uf_destino",
}

# Componentes do frete (opcionais). Cabeçalho normalizado -> categoria canônica.
# As categorias coincidem com as produzidas pelo parser de XML (cte_parser),
# garantindo que a análise de composição trate XML e Excel de forma idêntica.
COMPONENTES = {
    "frete peso": "Frete Peso",
    "frete-peso": "Frete Peso",
    "frete valor": "Frete Valor",
    "frete-valor": "Frete Valor",
    "ad valorem": "Frete Valor",
    "advalorem": "Frete Valor",
    "pedagio": "Pedágio",
    "pedágio": "Pedágio",
    "gris": "GRIS",
    "seguro": "Seguro",
    "ademe": "Ademe",
    "despacho": "Despacho",
    "outros": "Outros",
    "outras taxas": "Outros",
    "taxas": "Outros",
    # Granularização RN-73 (v6.7.0) — mesmas categorias novas do cte_parser.
    # Match exato por cabeçalho normalizado (sem heurística de substring, ao
    # contrário do cte_parser): variações de cabeçalho não listadas aqui
    # continuam caindo fora do dicionário (coluna ignorada), não em "Outros".
    "tde": "TDE",
    "tda": "TDA",
    "tda/tde": "Outros",  # item combinado — não decomponível (RN-73)
    "estadia": "Estadia",
}

OBRIGATORIAS = {
    "nota_fiscal", "data_emissao", "transportadora", "cidade_origem",
    "cidade_destino", "peso", "valor_mercadoria", "valor_frete",
}


class LinhaExcel:
    def __init__(self, dados: dict, componentes: Optional[dict] = None):
        # Célula vazia no Excel vira None (não ""): usar `or ""` evita que
        # str(None) vire a string literal "None" no campo.
        self.nota_fiscal: str = str(dados.get("nota_fiscal") or "").strip()
        self.cte: str = str(dados.get("cte") or "").strip()
        self.data_emissao: Optional[date] = _to_date(dados.get("data_emissao"))
        self.data_embarque: Optional[date] = _to_date(dados.get("data_embarque"))
        self.data_saida: Optional[date] = _to_date(dados.get("data_saida"))
        self.data_entrega: Optional[date] = _to_date(dados.get("data_entrega"))
        # Destinatário (cliente final) — dimensão CLIENTE_FINAL do DLG (RN-67/68).
        # Ambos opcionais: sem nenhum dos dois, o CT-e cai em "Cliente não
        # identificado" (mesmo tratamento já dado ao XML sem <dest>).
        self.destinatario_nome: str = str(dados.get("destinatario_nome", "")).strip()
        self.destinatario_cnpj: str = _so_digitos(dados.get("destinatario_cnpj"))
        # Tomador (opcional) — CNPJ da matriz/filial responsável pelo frete
        # (RN v6.15). Alimenta a dimensão FILIAL do DLG, igual ao XML. Sem
        # essa coluna, o CT-e cai em "Filial não identificada" (mesmo
        # tratamento já dado ao Destinatário ausente).
        self.tomador_cnpj: str = _so_digitos(dados.get("tomador_cnpj"))
        self.transportadora: str = str(dados.get("transportadora", "")).strip()
        self.cidade_origem: str = str(dados.get("cidade_origem", "")).strip()
        self.cidade_destino: str = str(dados.get("cidade_destino", "")).strip()
        self.uf_origem: str = str(dados.get("uf_origem", "")).strip().upper()[:2]
        self.uf_destino: str = str(dados.get("uf_destino", "")).strip().upper()[:2]
        self.peso: float = _to_float(dados.get("peso"))
        self.valor_mercadoria: float = _to_float(dados.get("valor_mercadoria"))
        self.valor_frete: float = _to_float(dados.get("valor_frete"))
        # Composição do frete: soma por categoria, apenas valores positivos.
        self.composicao: dict = {}
        for categoria, valor in (componentes or {}).items():
            v = _to_float(valor)
            if v > 0:
                self.composicao[categoria] = round(
                    self.composicao.get(categoria, 0.0) + v, 2
                )


def _normalizar(texto: str) -> str:
    import unicodedata
    s = "".join(
        c for c in unicodedata.normalize("NFKD", str(texto))
        if not unicodedata.combining(c)
    )
    return s.strip().lower()


def parse_excel(file_bytes: bytes) -> tuple[List[LinhaExcel], int]:
    from io import BytesIO

    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        raise ValueError("Planilha vazia.")

    # Mapeia índice de coluna -> campo canônico (escalares) e -> componente (frete)
    indice_para_campo: dict[int, str] = {}
    indice_para_componente: dict[int, str] = {}
    for idx, nome in enumerate(header):
        if nome is None:
            continue
        chave = _normalizar(nome)
        if chave in COLUNAS:
            indice_para_campo[idx] = COLUNAS[chave]
        elif chave in COMPONENTES:
            indice_para_componente[idx] = COMPONENTES[chave]

    encontradas = set(indice_para_campo.values())
    faltando = OBRIGATORIAS - encontradas
    if faltando:
        raise ValueError(
            "Colunas obrigatórias ausentes na planilha: "
            + ", ".join(sorted(faltando))
        )

    linhas: List[LinhaExcel] = []
    sem_identificador = 0
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        dados = {
            campo: row[idx]
            for idx, campo in indice_para_campo.items()
            if idx < len(row)
        }
        # Nota Fiscal e CT-e são identificadores alternativos: a linha só é
        # descartada quando NENHum dos dois está preenchido (ex.: planilha
        # que só informa o número do CT-e, sem coluna de NF).
        tem_nf = bool(str(dados.get("nota_fiscal") or "").strip())
        tem_cte = bool(str(dados.get("cte") or "").strip())
        if not tem_nf and not tem_cte:
            sem_identificador += 1
            continue
        componentes = {
            categoria: row[idx]
            for idx, categoria in indice_para_componente.items()
            if idx < len(row)
        }
        linhas.append(LinhaExcel(dados, componentes))

    wb.close()
    return linhas, sem_identificador


def _so_digitos(valor) -> str:
    # Excel pode devolver CNPJ como número (célula não formatada como texto);
    # nesse caso str(float) traria ".0" no final — normaliza para int antes.
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return "".join(c for c in str(valor or "") if c.isdigit())


def _to_float(valor) -> float:
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip().replace("R$", "").replace(" ", "")
    # trata formato brasileiro 1.234,56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_date(valor) -> Optional[date]:
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    s = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── Modelo de planilha para download (RF010 — apoio à importação) ─────────────
# Cabeçalhos amigáveis que o parser reconhece (mantém o modelo em sincronia
# com as colunas aceitas). Cada nome normaliza para uma chave de COLUNAS ou
# COMPONENTES — alterar aqui reflete no modelo baixado pelos usuários.
MODELO_COLUNAS = [
    ("Nota Fiscal", "12345", True),
    ("CT-e", "CTE-0001", False),
    ("Data Emissao", "01/03/2025", True),
    ("Data Embarque", "01/03/2025", False),
    ("Data Saida", "01/03/2025", False),
    ("Data Entrega", "05/03/2025", False),
    ("Transportadora", "Transportadora Exemplo Ltda", True),
    ("Destinatario", "Cliente Exemplo Ltda", False),
    ("Destinatario CNPJ", "12345678000199", False),
    ("CNPJ Tomador", "12345678000100", False),
    ("Cidade Origem", "Sao Paulo", True),
    ("UF Origem", "SP", False),
    ("Cidade Destino", "Rio de Janeiro", True),
    ("UF Destino", "RJ", False),
    ("Peso", "100", True),
    ("Valor Mercadoria", "5000,00", True),
    ("Valor Frete", "250,00", True),
    ("Frete Peso", "180,00", False),
    ("Frete Valor", "30,00", False),
    ("Pedagio", "20,00", False),
    ("GRIS", "10,00", False),
    ("Seguro", "5,00", False),
    ("Ademe", "0,00", False),
    ("Despacho", "0,00", False),
    ("Outros", "5,00", False),
]


def gerar_modelo_excel() -> bytes:
    """Gera um arquivo .xlsx modelo para importação, com cabeçalho, uma linha
    de exemplo e uma aba de instruções. As colunas coincidem exatamente com as
    reconhecidas por `parse_excel`."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Importação"

    cabecalho_fill = PatternFill("solid", fgColor="2D3561")
    obrig_font = Font(bold=True, color="FFFFFF")
    opc_font = Font(bold=True, color="FFE7A8")

    for col, (nome, exemplo, obrigatoria) in enumerate(MODELO_COLUNAS, start=1):
        c = ws.cell(row=1, column=col, value=nome)
        c.fill = cabecalho_fill
        c.font = obrig_font if obrigatoria else opc_font
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=2, column=col, value=exemplo)
        ws.column_dimensions[c.column_letter].width = max(14, len(nome) + 2)

    ws.freeze_panes = "A2"

    inst = wb.create_sheet("Instruções")
    linhas = [
        ("Modelo de importação — GD Diagnóstico Logístico", True),
        ("", False),
        ("Colunas em branco (cabeçalho azul/branco) são OBRIGATÓRIAS:", True),
        ("Nota Fiscal, Data Emissao, Transportadora, Cidade Origem,", False),
        ("Cidade Destino, Peso, Valor Mercadoria, Valor Frete.", False),
        ("", False),
        ("Data Emissao é a data de emissão do CT-e (mês/ano usado em todo", True),
        ("indicador e filtro por período) — não confundir com Data Embarque,", False),
        ("Data Saida ou Data Entrega, que são datas operacionais opcionais.", False),
        ("", False),
        ("Colunas em destaque dourado são OPCIONAIS:", True),
        ("CT-e, Data Embarque/Saida/Entrega, Destinatario, Destinatario CNPJ,", False),
        ("CNPJ Tomador, UFs e os componentes de frete (Frete Peso, Frete Valor,", False),
        ("Pedágio, GRIS, Seguro, Ademe, Despacho, Outros).", False),
        ("", False),
        ("Destinatario / Destinatario CNPJ identificam o cliente final da carga", True),
        ("e alimentam a dimensão Cliente do Diagnóstico e das Recomendações.", False),
        ("Sem preencher, o CT-e entra como \"Cliente não identificado\".", False),
        ("", False),
        ("CNPJ Tomador (ou apenas \"Tomador\"/\"Filial\") identifica a matriz/filial", True),
        ("responsável pelo frete e alimenta a dimensão Filial do Diagnóstico.", False),
        ("Aceita só os números do CNPJ ou um texto com o CNPJ embutido", False),
        ("(ex: \"03822909000113 - Razão Social\") — os dígitos são extraídos", False),
        ("automaticamente. Deve ser o CNPJ de uma filial (ou da matriz) já", False),
        ("cadastrada em Empresas e Filiais. Sem preencher, o CT-e entra como", False),
        ("\"Filial não identificada\".", False),
        ("", False),
        ("Observações:", True),
        ("- A ordem das colunas é livre; os nomes (com ou sem acento) são reconhecidos.", False),
        ("- Valores em R$ aceitam o formato brasileiro (1.234,56).", False),
        ("- Datas aceitam dd/mm/aaaa.", False),
        ("- Apague a linha de exemplo (linha 2) antes de importar.", False),
        ("- Os componentes de frete alimentam a análise de divergência por componente.", False),
    ]
    for i, (txt, negrito) in enumerate(linhas, start=1):
        cel = inst.cell(row=i, column=1, value=txt)
        if negrito:
            cel.font = Font(bold=True)
    inst.column_dimensions["A"].width = 80

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
