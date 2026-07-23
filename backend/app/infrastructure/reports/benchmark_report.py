"""Relatórios consolidados do Módulo Benchmark Logístico (seção 14).

Gera, em PDF, HTML e Excel, um relatório único cobrindo Benchmark Nacional,
Regional, Transportadoras e Potencial de Economia — reaproveitando os
estilos e a paleta GD Conecta já definidos nos relatórios de diagnóstico.
"""
from html import escape
from io import BytesIO

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.application.dtos import (
    BenchmarkNacional,
    PotencialEconomia,
)
from app.infrastructure.reports.excel_report import _autofit, _style_header, _title_font
from app.infrastructure.reports.pdf_report import _br, _estilos, _tabela
from app.infrastructure.reports._html_theme import (
    DANGER, DANGER_BG, OK, OK_BG, WARN, WARN_BG,
    chip, moeda, numero, pct, render_documento,
)


def _rotulo_macro(m: str) -> str:
    return {
        "NORTE": "Norte", "NORDESTE": "Nordeste", "CENTRO_OESTE": "Centro-Oeste",
        "SUDESTE": "Sudeste", "SUL": "Sul",
    }.get(m, m)


# ============================ PDF ============================
def gerar_benchmark_pdf(
    empresa_nome: str,
    nacional: BenchmarkNacional,
    regional: list,
    transportadoras: list,
    economia: PotencialEconomia,
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=1.6 * cm, bottomMargin=1.6 * cm,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
        title=f"Benchmark Logístico — {empresa_nome}",
    )
    s = _estilos()
    el = []
    el.append(Paragraph("Relatório de Benchmark Logístico", s["GDTitulo"]))
    el.append(Paragraph(f"{empresa_nome} — comparativo com o mercado", s["GDInfo"]))
    el.append(Spacer(1, 6))

    # --- Benchmark Nacional ---
    el.append(Paragraph("1. Benchmark Nacional", s["GDSub"]))
    el.append(_tabela([
        ["Indicador", "Empresa", "Bench. médio", "Desvio", "Classificação"],
        ["Custo por kg", _br(nacional.frete_kg.valor, 4), _br(nacional.frete_kg.benchmark_medio, 4),
         f"{_br(nacional.frete_kg.desvio_pct, 1)}%", nacional.frete_kg.classificacao],
        ["% Frete s/ mercadoria", f"{_br(nacional.frete_pct.valor, 2)}%",
         f"{_br(nacional.frete_pct.benchmark_medio, 2)}%",
         f"{_br(nacional.frete_pct.desvio_pct, 1)}%", nacional.frete_pct.classificacao],
    ]))

    # --- Benchmark Regional ---
    el.append(Paragraph("2. Benchmark Regional", s["GDSub"]))
    linhas = [["Região", "R$/kg", "Bench.", "Desvio", "Classif.", "% Frete"]]
    for r in regional:
        linhas.append([
            _rotulo_macro(r.macro_regiao), _br(r.frete_kg.valor, 4),
            _br(r.frete_kg.benchmark_medio, 4), f"{_br(r.frete_kg.desvio_pct, 1)}%",
            r.frete_kg.classificacao, f"{_br(r.frete_pct.valor, 2)}%",
        ])
    el.append(_tabela(linhas))

    # --- Transportadoras ---
    el.append(Paragraph("3. Benchmark por Transportadora", s["GDSub"]))
    linhas = [["Transportadora", "Frete total", "R$/kg", "Desvio", "Nível", "Classif."]]
    for t in transportadoras[:20]:
        linhas.append([
            t.nome[:28], _br(t.frete_total), _br(t.frete_kg.valor, 4),
            f"{_br(t.frete_kg.desvio_pct, 1)}%", t.nivel_custo, t.frete_kg.classificacao,
        ])
    el.append(_tabela(linhas))

    # --- Potencial de Economia ---
    el.append(Paragraph("4. Potencial de Economia", s["GDSub"]))
    el.append(_tabela([
        ["Métrica", "Valor"],
        ["Economia no período", f"R$ {_br(economia.economia_total)}"],
        ["% sobre o frete", f"{_br(economia.economia_pct, 1)}%"],
        ["Projeção mensal", f"R$ {_br(economia.proj_mensal)}"],
        ["Projeção trimestral", f"R$ {_br(economia.proj_trimestral)}"],
        ["Projeção semestral", f"R$ {_br(economia.proj_semestral)}"],
        ["Projeção anual", f"R$ {_br(economia.proj_anual)}"],
    ], col_widths=[9 * cm, 7 * cm]))

    el.append(Spacer(1, 10))
    el.append(Paragraph(
        "GD Conecta — Governança de Frete · Análise de Dados · Torre de Controle",
        s["GDInfo"],
    ))
    doc.build(el)
    return buf.getvalue()


# ============================ HTML ============================
def _classe_bench(classificacao: str) -> str:
    c = (classificacao or "").lower()
    if "crítico" in c or "critico" in c:
        return "critico"
    if "atenção" in c or "atencao" in c:
        return "atencao"
    if "excelente" in c or "bom" in c:
        return "eficiente"
    return "neutro"


def _saude_geral_benchmark(nacional: BenchmarkNacional) -> tuple:
    classes = {_classe_bench(nacional.frete_kg.classificacao), _classe_bench(nacional.frete_pct.classificacao)}
    if "critico" in classes:
        return ("Crítico", DANGER, DANGER_BG)
    if "atencao" in classes:
        return ("Atenção", WARN, WARN_BG)
    return ("Saudável", OK, OK_BG)


def _resumo_executivo_benchmark(nacional: BenchmarkNacional, economia: PotencialEconomia) -> str:
    desvio = nacional.frete_kg.desvio_pct
    direcao = "acima" if desvio > 0 else "abaixo"
    partes = [f"Custo por kg {numero(abs(desvio), 1)}% {direcao} do benchmark de mercado"]
    if economia.proj_anual:
        partes.append(f"economia potencial projetada em {moeda(economia.proj_anual)}/ano")
    return escape(" — ".join(partes)) + "."


def _recomendacoes_benchmark(regional: list, transportadoras: list, economia: PotencialEconomia) -> str:
    """Insights derivados diretamente dos mesmos DTOs do relatório — o
    Benchmark não tem recomendações persistidas como o Diagnóstico, então
    são calculadas aqui a partir dos maiores desvios frente ao benchmark."""
    cards = []
    piores_regioes = [r for r in regional if r.frete_kg.desvio_pct > 0]
    if piores_regioes:
        pior = max(piores_regioes, key=lambda r: r.frete_kg.desvio_pct)
        cards.append((
            "atencao",
            f"Região {_rotulo_macro(pior.macro_regiao)} acima do benchmark",
            f"O custo por kg nesta região está {numero(pior.frete_kg.desvio_pct, 1)}% acima da referência de "
            "mercado — avalie renegociação de frete ou revisão de rotas nesta região.",
        ))
    piores_transp = [t for t in transportadoras if t.frete_kg.desvio_pct > 0]
    if piores_transp:
        pior_t = max(piores_transp, key=lambda t: t.frete_kg.desvio_pct)
        cards.append((
            "atencao",
            f"Transportadora {pior_t.nome} com custo elevado",
            f"R$/kg {numero(pior_t.frete_kg.desvio_pct, 1)}% acima do benchmark, com participação de "
            f"{numero(pior_t.participacao_pct, 1)}% no frete total — priorize esta transportadora numa "
            "próxima negociação ou BID.",
        ))
    if economia.economia_total > 0:
        cards.append((
            "neutro",
            "Potencial de economia identificado",
            "Ajustando os indicadores acima ao benchmark de mercado, a economia potencial no período é de "
            f"{moeda(economia.economia_total)} ({numero(economia.economia_pct, 1)}% do frete total).",
        ))
    if not cards:
        return '<p class="vazio">Nenhum desvio relevante frente ao benchmark de mercado neste período.</p>'
    return "".join(
        f'<div class="recomendacao {classe}"><div class="linha-topo"><h3>{escape(titulo)}</h3></div>'
        f"<p>{escape(descricao)}</p></div>"
        for classe, titulo, descricao in cards
    )


def gerar_benchmark_html(
    empresa_nome: str,
    nacional: BenchmarkNacional,
    regional: list,
    transportadoras: list,
    economia: PotencialEconomia,
    gerado_em: str,
) -> bytes:
    periodo = "Todo o período"
    if nacional.periodo_inicio or nacional.periodo_fim:
        periodo = f"{nacional.periodo_inicio or '...'} a {nacional.periodo_fim or '...'}"

    linhas_regiao = "".join(
        "<tr>"
        f"<td>{escape(_rotulo_macro(r.macro_regiao))}</td>"
        f"<td>{moeda(r.frete_total)}</td>"
        f"<td>{numero(r.frete_kg.valor, 4)}</td>"
        f"<td>{numero(r.frete_kg.benchmark_medio, 4)}</td>"
        f'<td class="{"desvio-pos" if r.frete_kg.desvio_pct > 0 else "desvio-neg"}">'
        f'{"+" if r.frete_kg.desvio_pct > 0 else ""}{numero(r.frete_kg.desvio_pct, 1)}%</td>'
        f'<td style="text-align:center">{chip(r.frete_kg.classificacao, _classe_bench(r.frete_kg.classificacao))}</td>'
        "</tr>"
        for r in regional
    ) or '<tr><td colspan="6" class="vazio">Sem dados regionais.</td></tr>'

    top_transp = sorted(transportadoras, key=lambda t: t.frete_total, reverse=True)[:20]
    linhas_transp = "".join(
        "<tr>"
        f"<td>{escape(t.nome)}</td>"
        f"<td>{moeda(t.frete_total)}</td>"
        f"<td>{numero(t.frete_kg.valor, 4)}</td>"
        f'<td class="{"desvio-pos" if t.frete_kg.desvio_pct > 0 else "desvio-neg"}">'
        f'{"+" if t.frete_kg.desvio_pct > 0 else ""}{numero(t.frete_kg.desvio_pct, 1)}%</td>'
        f"<td>{escape(t.nivel_custo)}</td>"
        f'<td style="text-align:center">{chip(t.frete_kg.classificacao, _classe_bench(t.frete_kg.classificacao))}</td>'
        "</tr>"
        for t in top_transp
    ) or '<tr><td colspan="6" class="vazio">Sem dados de transportadoras.</td></tr>'

    linhas_economia_regiao = "".join(
        "<tr>"
        f"<td>{escape(_rotulo_macro(r.macro_regiao))}</td>"
        f"<td>{numero(r.frete_rs_kg, 4)}</td>"
        f"<td>{moeda(r.economia)}</td>"
        "</tr>"
        for r in economia.por_regiao
    ) or '<tr><td colspan="3" class="vazio">Sem dados de economia por região.</td></tr>'

    bloco_recomendacoes = _recomendacoes_benchmark(regional, transportadoras, economia)

    corpo_html = f"""
      <section class="bloco">
        <div class="titulo-secao"><h2>Benchmark Nacional</h2></div>
        <div class="grade-kpi tabular">
          <div class="kpi"><span class="rotulo">Custo por kg</span><div class="valor">{numero(nacional.frete_kg.valor, 4)}</div>
            <div class="legenda {"desvio-pos" if nacional.frete_kg.desvio_pct > 0 else "desvio-neg"}">{numero(nacional.frete_kg.desvio_pct, 1)}% vs. benchmark ({numero(nacional.frete_kg.benchmark_medio, 4)})</div>
          </div>
          <div class="kpi"><span class="rotulo">Classificação R$/kg</span><div class="valor" style="font-size:15px">{chip(nacional.frete_kg.classificacao, _classe_bench(nacional.frete_kg.classificacao))}</div></div>
          <div class="kpi"><span class="rotulo">% Frete / mercadoria</span><div class="valor">{pct(nacional.frete_pct.valor)}</div>
            <div class="legenda">benchmark: {pct(nacional.frete_pct.benchmark_medio)}</div>
          </div>
          <div class="kpi"><span class="rotulo">Classificação % Frete</span><div class="valor" style="font-size:15px">{chip(nacional.frete_pct.classificacao, _classe_bench(nacional.frete_pct.classificacao))}</div></div>
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Benchmark Regional</h2></div>
        <div class="tabela-wrap">
          <table class="tabular">
            <thead><tr><th>Região</th><th>Frete total</th><th>R$/kg</th><th>Bench.</th><th>Desvio</th><th style="text-align:center">Classif.</th></tr></thead>
            <tbody>{linhas_regiao}</tbody>
          </table>
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Benchmark por Transportadora</h2><span class="nota">top 20 por frete total</span></div>
        <div class="tabela-wrap">
          <table class="tabular">
            <thead><tr><th>Transportadora</th><th>Frete total</th><th>R$/kg</th><th>Desvio</th><th>Nível custo</th><th style="text-align:center">Classif.</th></tr></thead>
            <tbody>{linhas_transp}</tbody>
          </table>
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Potencial de Economia</h2></div>
        <div class="grade-kpi tabular">
          <div class="kpi"><span class="rotulo">Economia no período</span><div class="valor">{moeda(economia.economia_total)}</div><div class="legenda">{numero(economia.economia_pct, 1)}% do frete</div></div>
          <div class="kpi"><span class="rotulo">Projeção mensal</span><div class="valor">{moeda(economia.proj_mensal)}</div></div>
          <div class="kpi"><span class="rotulo">Projeção trimestral</span><div class="valor">{moeda(economia.proj_trimestral)}</div></div>
          <div class="kpi"><span class="rotulo">Projeção semestral</span><div class="valor">{moeda(economia.proj_semestral)}</div></div>
          <div class="kpi"><span class="rotulo">Projeção anual</span><div class="valor">{moeda(economia.proj_anual)}</div></div>
        </div>
        <div class="tabela-wrap" style="margin-top:16px">
          <table class="tabular">
            <thead><tr><th>Região</th><th>R$/kg atual</th><th>Economia</th></tr></thead>
            <tbody>{linhas_economia_regiao}</tbody>
          </table>
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Recomendações</h2></div>
        <p class="descricao-secao">Oportunidades identificadas a partir do comparativo com o benchmark de mercado.</p>
        {bloco_recomendacoes}
      </section>
"""

    return render_documento(
        titulo_pagina=f"Benchmark de Mercado — {escape(empresa_nome)}",
        titulo_relatorio="Relatório de Benchmark de Mercado",
        meta_items=[
            f"<span>Empresa: <b>{escape(empresa_nome)}</b></span>",
            f"<span>Período: <b>{escape(periodo)}</b></span>",
            f"<span>Gerado em: <b>{escape(gerado_em)}</b></span>",
        ],
        corpo_html=corpo_html,
        selo_saude=_saude_geral_benchmark(nacional),
        resumo_executivo=_resumo_executivo_benchmark(nacional, economia),
        rodape_nota="GD Conecta — Governança de Frete · Benchmark de Mercado.",
    )


# ============================ Excel ============================
def gerar_benchmark_excel(
    empresa_nome: str,
    nacional: BenchmarkNacional,
    regional: list,
    transportadoras: list,
    economia: PotencialEconomia,
) -> bytes:
    from openpyxl import Workbook

    periodo = "Todo o período"
    if nacional.periodo_inicio or nacional.periodo_fim:
        periodo = f"{nacional.periodo_inicio or '...'} a {nacional.periodo_fim or '...'}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Nacional"
    ws.cell(row=1, column=1, value=f"Benchmark Logístico — {empresa_nome}").font = _title_font
    ws.cell(row=2, column=1, value=f"Período: {periodo}")
    ws.append([])
    ws.append(["Indicador", "Empresa", "Bench. médio", "Desvio %", "Classificação"])
    _style_header(ws, ws.max_row, 5)
    ws.append([
        "Custo por kg", round(nacional.frete_kg.valor, 4), round(nacional.frete_kg.benchmark_medio, 4),
        round(nacional.frete_kg.desvio_pct, 1), nacional.frete_kg.classificacao,
    ])
    ws.append([
        "% Frete / mercadoria", round(nacional.frete_pct.valor, 2), round(nacional.frete_pct.benchmark_medio, 2),
        round(nacional.frete_pct.desvio_pct, 1), nacional.frete_pct.classificacao,
    ])
    _autofit(ws)

    ws_reg = wb.create_sheet("Regional")
    ws_reg.append(["Região", "Frete total (R$)", "R$/kg", "Bench. R$/kg", "Desvio %", "Classif.", "% Frete"])
    _style_header(ws_reg, ws_reg.max_row, 7)
    for r in regional:
        ws_reg.append([
            _rotulo_macro(r.macro_regiao), round(r.frete_total, 2), round(r.frete_kg.valor, 4),
            round(r.frete_kg.benchmark_medio, 4), round(r.frete_kg.desvio_pct, 1),
            r.frete_kg.classificacao, round(r.frete_pct.valor, 2),
        ])
    _autofit(ws_reg)

    ws_transp = wb.create_sheet("Transportadoras")
    ws_transp.append(["Transportadora", "Frete total (R$)", "R$/kg", "Desvio %", "Nível custo", "Classificação", "Participação %"])
    _style_header(ws_transp, ws_transp.max_row, 7)
    for t in sorted(transportadoras, key=lambda t: t.frete_total, reverse=True):
        ws_transp.append([
            t.nome, round(t.frete_total, 2), round(t.frete_kg.valor, 4),
            round(t.frete_kg.desvio_pct, 1), t.nivel_custo, t.frete_kg.classificacao,
            round(t.participacao_pct, 1),
        ])
    _autofit(ws_transp)

    ws_eco = wb.create_sheet("Economia")
    ws_eco.append(["Métrica", "Valor"])
    _style_header(ws_eco, ws_eco.max_row, 2)
    ws_eco.append(["Economia no período (R$)", round(economia.economia_total, 2)])
    ws_eco.append(["% sobre o frete", round(economia.economia_pct, 1)])
    ws_eco.append(["Projeção mensal (R$)", round(economia.proj_mensal, 2)])
    ws_eco.append(["Projeção trimestral (R$)", round(economia.proj_trimestral, 2)])
    ws_eco.append(["Projeção semestral (R$)", round(economia.proj_semestral, 2)])
    ws_eco.append(["Projeção anual (R$)", round(economia.proj_anual, 2)])
    ws_eco.append([])
    ws_eco.append(["Região", "R$/kg atual", "Economia (R$)"])
    _style_header(ws_eco, ws_eco.max_row, 3)
    for reg in economia.por_regiao:
        ws_eco.append([_rotulo_macro(reg.macro_regiao), round(reg.frete_rs_kg, 4), round(reg.economia, 2)])
    _autofit(ws_eco)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
