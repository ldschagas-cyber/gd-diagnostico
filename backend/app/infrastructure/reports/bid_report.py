"""Gerador de relatórios do módulo BID — PDF e Excel.

Tipos disponíveis:
  - executivo: visão geral do BID
  - comparativo: proposta vs. atual por grupo
  - ranking: score de transportadoras
  - economia: projeções de saving
  - resultado: BID encerrado com vencedoras

Pacote de Cotação (formato especial): sem valores pagos atualmente.
"""
from html import escape
from io import BytesIO

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.infrastructure.reports.excel_report import _autofit, _style_header, _title_font
from app.infrastructure.reports.pdf_report import _br, _estilos, _tabela
from app.infrastructure.reports._html_theme import (
    BLUE, DANGER, DANGER_BG, IVORY_DEEP, OK, OK_BG, SLATE, WARN, WARN_BG,
    chip, moeda, numero, render_documento,
)


def _rotulo(tipo: str) -> str:
    return {
        "REGIAO": "Região", "UF": "UF", "FILIAL": "Filial",
        "TRANSPORTADORA": "Transportadora", "FAIXA_PESO": "Faixa de Peso",
    }.get(tipo, tipo)


# ═══════════════════════════════════════════ PDF ════════════════════════════

def gerar_bid_pdf(tipo: str, empresa_nome: str, bid, dados: dict) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        title=f"BID {bid.nome} — {tipo.title()}",
    )
    s = _estilos()
    el = []

    el.append(Paragraph(f"BID de Frete — {tipo.replace('_', ' ').title()}", s["GDTitulo"]))
    el.append(Paragraph(f"{empresa_nome}  ·  {bid.nome}", s["GDInfo"]))
    el.append(Paragraph(f"Status: {bid.status.value}  ·  Período: {bid.periodo_analise_inicio} → {bid.periodo_analise_fim}", s["GDInfo"]))
    el.append(Spacer(1, 8))

    if tipo == "executivo":
        el += _secao_executivo(bid, dados, s)
    elif tipo == "comparativo":
        el += _secao_comparativo(dados, s)
    elif tipo == "ranking":
        el += _secao_ranking(dados, s)
    elif tipo == "economia":
        el += _secao_economia(dados, s)
    elif tipo == "resultado":
        el += _secao_resultado(dados, s)

    el.append(Spacer(1, 12))
    el.append(Paragraph("GD Conecta — Governança de Frete · Torre de Controle Logístico", s["GDInfo"]))
    doc.build(el)
    return buf.getvalue()


def _secao_executivo(bid, dados, s):
    el = []
    el.append(Paragraph("1. Escopo do BID", s["GDSub"]))
    escopos = dados.get("escopos", [])
    linhas = [["Agrupamento", "Grupo", "Embarques", "Peso (kg)", "Frete (R$)", "R$/kg"]]
    for e in escopos:
        linhas.append([
            _rotulo(e.tipo_agrupamento.value), e.valor_grupo,
            str(e.qtd_embarques), _br(e.peso_total, 0),
            _br(e.valor_frete_total), _br(e.frete_rs_kg, 4),
        ])
    el.append(_tabela(linhas))
    el.append(Spacer(1, 8))
    el.append(Paragraph("2. Transportadoras Convidadas", s["GDSub"]))
    bts = dados.get("bid_transportadoras", [])
    transp_nomes = dados.get("transp_nomes", {})
    linhas = [["Transportadora", "Status"]]
    for bt in bts:
        linhas.append([transp_nomes.get(bt.transportadora_id, str(bt.transportadora_id)), bt.status.value])
    el.append(_tabela(linhas))
    return el


def _secao_comparativo(dados, s):
    el = []
    el.append(Paragraph("Comparativo de Propostas", s["GDSub"]))
    comparativo = dados.get("comparativo", [])
    linhas = [["Grupo", "Atual R$/kg", "Transportadora", "Proposta R$/kg", "Economia R$", "Economia %"]]
    for linha in comparativo:
        for p in linha.propostas:
            linhas.append([
                linha.valor_grupo, _br(linha.frete_atual_rs_kg, 4),
                p["transportadora"], _br(p["rs_kg"], 4),
                _br(p["economia_rs"]), f"{_br(p['economia_pct'], 1)}%",
            ])
    el.append(_tabela(linhas))
    return el


def _secao_ranking(dados, s):
    el = []
    el.append(Paragraph("Ranking de Transportadoras", s["GDSub"]))
    scores = dados.get("scores", [])
    linhas = [["#", "Transportadora", "Score", "Classificação", "R$/kg médio", "Prazo médio"]]
    for i, sc in enumerate(scores, 1):
        linhas.append([
            str(i), sc.nome, str(sc.score), sc.classificacao,
            _br(sc.proposta_media_rs_kg, 4), f"{_br(sc.prazo_medio, 0)} dias",
        ])
    el.append(_tabela(linhas))
    return el


def _secao_economia(dados, s):
    el = []
    eco = dados.get("economia")
    if eco:
        el.append(Paragraph("Potencial de Economia", s["GDSub"]))
        el.append(_tabela([
            ["Métrica", "Valor"],
            ["Frete atual total", f"R$ {_br(eco.frete_atual_total)}"],
            ["Economia no período", f"R$ {_br(eco.economia_total_periodo)}"],
            ["% de redução", f"{_br(eco.economia_pct, 1)}%"],
            ["Ritmo mensal", f"R$ {_br(eco.ritmo_mensal)}"],
            ["Projeção trimestral", f"R$ {_br(eco.proj_trimestral)}"],
            ["Projeção semestral", f"R$ {_br(eco.proj_semestral)}"],
            ["Projeção anual", f"R$ {_br(eco.proj_anual)}"],
        ], col_widths=[10 * cm, 7 * cm]))
        el.append(Spacer(1, 6))
        el.append(Paragraph("Por grupo", s["GDSub"]))
        linhas = [["Grupo", "R$/kg atual", "Melhor proposta", "Economia R$"]]
        for g in eco.por_grupo:
            linhas.append([
                g["valor_grupo"], _br(g["frete_atual_rs_kg"], 4),
                _br(g["melhor_proposta_rs_kg"], 4), _br(g["economia"]),
            ])
        el.append(_tabela(linhas))
    return el


def _secao_resultado(dados, s):
    el = _secao_comparativo(dados, s)
    el += _secao_ranking(dados, s)
    el += _secao_economia(dados, s)
    return el


# ══════════════════════════════════════════ Excel ═══════════════════════════

def gerar_bid_excel(tipo: str, empresa_nome: str, bid, dados: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = tipo.title()

    ws.cell(row=1, column=1, value=f"BID: {bid.nome} — {empresa_nome}").font = _title_font
    ws.cell(row=2, column=1, value=f"Status: {bid.status.value} | Período: {bid.periodo_analise_inicio} → {bid.periodo_analise_fim}")
    ws.append([])

    if tipo == "comparativo":
        _aba_comparativo(ws, dados)
    elif tipo == "ranking":
        _aba_ranking(ws, dados)
    elif tipo == "economia":
        _aba_economia(ws, dados, wb)
    elif tipo == "resultado":
        _aba_comparativo(ws, dados)
        ws2 = wb.create_sheet("Score")
        _aba_ranking(ws2, dados)
        ws3 = wb.create_sheet("Economia")
        _aba_economia(ws3, dados, wb)

    _autofit(ws)
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _aba_comparativo(ws, dados):
    ws.append(["Grupo", "Atual R$/kg", "Transportadora", "Proposta R$/kg", "Economia R$", "Economia %"])
    _style_header(ws, ws.max_row, 6)
    for linha in dados.get("comparativo", []):
        for p in linha.propostas:
            ws.append([linha.valor_grupo, round(linha.frete_atual_rs_kg, 4),
                       p["transportadora"], round(p["rs_kg"], 4),
                       round(p["economia_rs"], 2), round(p["economia_pct"], 1)])


def _aba_ranking(ws, dados):
    ws.append(["#", "Transportadora", "Score", "Classificação", "R$/kg médio", "Prazo médio", "Cobertura %"])
    _style_header(ws, ws.max_row, 7)
    for i, sc in enumerate(dados.get("scores", []), 1):
        ws.append([i, sc.nome, sc.score, sc.classificacao,
                   round(sc.proposta_media_rs_kg, 4), round(sc.prazo_medio, 0), round(sc.cobertura_media, 1)])


def gerar_bid_executivo_excel(empresa_nome: str, bid, dados: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Escopo"
    ws.cell(row=1, column=1, value=f"BID: {bid.nome} — {empresa_nome}").font = _title_font
    ws.cell(row=2, column=1, value=f"Status: {bid.status.value} | Período: {bid.periodo_analise_inicio} → {bid.periodo_analise_fim}")
    ws.append([])
    ws.append(["Agrupamento", "Grupo", "Embarques", "Peso (kg)", "Frete (R$)", "R$/kg"])
    _style_header(ws, ws.max_row, 6)
    for e in dados.get("escopos", []):
        ws.append([
            _rotulo(e.tipo_agrupamento.value), e.valor_grupo, e.qtd_embarques,
            round(e.peso_total, 0), round(e.valor_frete_total, 2), round(e.frete_rs_kg, 4),
        ])
    _autofit(ws)

    transp_nomes = dados.get("transp_nomes", {})
    ws_transp = wb.create_sheet("Transportadoras")
    ws_transp.append(["Transportadora", "Status"])
    _style_header(ws_transp, ws_transp.max_row, 2)
    for bt in dados.get("bid_transportadoras", []):
        ws_transp.append([transp_nomes.get(bt.transportadora_id, str(bt.transportadora_id)), bt.status.value])
    _autofit(ws_transp)

    ws_rank = wb.create_sheet("Ranking")
    _aba_ranking(ws_rank, dados)
    _autofit(ws_rank)

    ws_eco = wb.create_sheet("Economia")
    _aba_economia(ws_eco, dados)
    _autofit(ws_eco)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _aba_economia(ws, dados, wb=None):
    eco = dados.get("economia")
    if not eco:
        return
    ws.append(["Métrica", "Valor (R$)"])
    _style_header(ws, ws.max_row, 2)
    ws.append(["Frete atual total", round(eco.frete_atual_total, 2)])
    ws.append(["Economia no período", round(eco.economia_total_periodo, 2)])
    ws.append(["% redução", round(eco.economia_pct, 1)])
    ws.append(["Ritmo mensal", round(eco.ritmo_mensal, 2)])
    ws.append(["Projeção trimestral", round(eco.proj_trimestral, 2)])
    ws.append(["Projeção semestral", round(eco.proj_semestral, 2)])
    ws.append(["Projeção anual", round(eco.proj_anual, 2)])
    ws.append([])
    ws.append(["Grupo", "R$/kg atual", "Melhor proposta", "Economia R$"])
    _style_header(ws, ws.max_row, 4)
    for g in eco.por_grupo:
        ws.append([g["valor_grupo"], round(g["frete_atual_rs_kg"], 4),
                   round(g["melhor_proposta_rs_kg"], 4), round(g["economia"], 2)])


# ══════════════════════════════════════ Pacote de Cotação ════════════════════

def gerar_pacote_cotacao_pdf(empresa_nome: str, bid, escopos: list) -> bytes:
    """Documento enviado às transportadoras. SEM valores de frete atuais."""
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                             leftMargin=1.5*cm, rightMargin=1.5*cm)
    s = _estilos()
    el = []
    el.append(Paragraph("Pacote de Cotação — BID de Frete", s["GDTitulo"]))
    el.append(Paragraph(f"{empresa_nome}  ·  {bid.nome}", s["GDInfo"]))
    if bid.data_encerramento:
        el.append(Paragraph(f"Prazo para propostas: {bid.data_encerramento}", s["GDInfo"]))
    el.append(Spacer(1, 8))
    if bid.objetivo:
        el.append(Paragraph("Objetivo", s["GDSub"]))
        el.append(Paragraph(bid.objetivo, s["GDInfo"]))
        el.append(Spacer(1, 6))
    el.append(Paragraph("Perfil Operacional (volume histórico)", s["GDSub"]))
    linhas = [["Grupo", "Tipo", "Embarques", "Peso total (kg)"]]
    for e in escopos:
        linhas.append([e.valor_grupo, _rotulo(e.tipo_agrupamento.value),
                       str(e.qtd_embarques), _br(e.peso_total, 0)])
    el.append(_tabela(linhas))
    el.append(Spacer(1, 10))
    el.append(Paragraph(
        "Por favor, envie sua proposta preenchendo: R$/kg, frete mínimo, prazo e cobertura por grupo.",
        s["GDInfo"],
    ))
    doc.build(el)
    return buf.getvalue()


# ═══════════════════════════════ HTML (executivo) ════════════════════════════

_ROTULO_STATUS_BID = {
    "RASCUNHO": "Rascunho", "ABERTO": "Aberto", "EM_COTACAO": "Em Cotação",
    "ENCERRADO": "Encerrado", "CANCELADO": "Cancelado",
}
_SELO_STATUS_BID = {
    "RASCUNHO": (SLATE, IVORY_DEEP),
    "ABERTO": (BLUE, "#E3F2FD"),
    "EM_COTACAO": (WARN, WARN_BG),
    "ENCERRADO": (OK, OK_BG),
    "CANCELADO": (DANGER, DANGER_BG),
}
_CLASSE_STATUS_TRANSP = {
    "VENCEDORA": "eficiente", "PROPOSTA_RECEBIDA": "eficiente",
    "PARTICIPANDO": "atencao", "CONVIDADA": "neutro",
    "DESISTIU": "critico", "DESCLASSIFICADA": "critico",
}
_PROXIMO_PASSO_BID = {
    "RASCUNHO": "BID ainda em rascunho — complete o escopo e convide transportadoras para iniciar a cotação.",
    "ABERTO": "BID em aberto — aguarde mais propostas antes de decidir, ou avance para contratação se o prazo estiver próximo do fim.",
    "EM_COTACAO": "Propostas em cotação — avalie o ranking acima e avance para a formalização com as transportadoras mais bem posicionadas.",
    "ENCERRADO": "BID encerrado — formalize a contratação das transportadoras vencedoras e monitore a economia contratada.",
    "CANCELADO": "BID cancelado — nenhuma ação necessária.",
}


def _classe_score(score: float) -> str:
    if score >= 80:
        return "eficiente"
    if score >= 65:
        return "neutro"
    if score >= 50:
        return "atencao"
    return "critico"


def _resumo_executivo_bid(escopos: list, scores: list, eco) -> str:
    partes = []
    if escopos:
        partes.append(f"{len(escopos)} grupo(s) avaliados")
    if scores:
        partes.append(f"{len(scores)} transportadora(s) rankeada(s)")
    if eco and eco.economia_total_periodo:
        partes.append(f"economia potencial de {moeda(eco.economia_total_periodo)} ({numero(eco.economia_pct, 1)}%)")
    return (escape(" — ".join(partes)) + ".") if partes else ""


def _recomendacoes_bid(scores: list, eco, status_val: str) -> str:
    """Insights derivados do ranking e da economia já calculados pelo
    EconomiaUseCase — não recalcula nada, só interpreta o resultado."""
    cards = []
    if scores:
        melhor = scores[0]
        cards.append((
            "eficiente",
            f"Transportadora recomendada: {melhor.nome}",
            f"Melhor colocada no ranking com score {numero(melhor.score, 0)} ({melhor.classificacao}), "
            f"R$/kg médio de {numero(melhor.proposta_media_rs_kg, 4)} e prazo médio de "
            f"{numero(melhor.prazo_medio, 0)} dias.",
        ))
    if eco and eco.economia_total_periodo:
        cards.append((
            "neutro",
            "Economia potencial identificada",
            f"Contratando as melhores propostas por grupo, a economia estimada é de "
            f"{moeda(eco.economia_total_periodo)} ({numero(eco.economia_pct, 1)}% de redução) no período analisado.",
        ))
    proximo_passo = _PROXIMO_PASSO_BID.get(status_val)
    if proximo_passo:
        cards.append(("neutro", "Próximo passo sugerido", proximo_passo))
    if not cards:
        return '<p class="vazio">Sem dados suficientes para recomendações neste momento.</p>'
    return "".join(
        f'<div class="recomendacao {classe}"><div class="linha-topo"><h3>{escape(titulo)}</h3></div>'
        f"<p>{escape(descricao)}</p></div>"
        for classe, titulo, descricao in cards
    )


def gerar_bid_executivo_html(empresa_nome: str, bid, dados: dict, gerado_em: str) -> bytes:
    status_val = bid.status.value
    rotulo_status = _ROTULO_STATUS_BID.get(status_val, status_val)
    cor, fundo = _SELO_STATUS_BID.get(status_val, (SLATE, IVORY_DEEP))

    escopos = dados.get("escopos", [])
    bts = dados.get("bid_transportadoras", [])
    transp_nomes = dados.get("transp_nomes", {})
    scores = sorted(dados.get("scores", []), key=lambda s: s.score, reverse=True)
    eco = dados.get("economia")

    linhas_escopo = "".join(
        "<tr>"
        f"<td>{escape(_rotulo(e.tipo_agrupamento.value))}</td>"
        f"<td>{escape(str(e.valor_grupo))}</td>"
        f"<td>{e.qtd_embarques}</td>"
        f"<td>{numero(e.peso_total, 0)}</td>"
        f"<td>{moeda(e.valor_frete_total)}</td>"
        f"<td>{numero(e.frete_rs_kg, 4)}</td>"
        "</tr>"
        for e in escopos
    ) or '<tr><td colspan="6" class="vazio">Sem escopo definido.</td></tr>'

    linhas_transp = "".join(
        "<tr>"
        f"<td>{escape(str(transp_nomes.get(bt.transportadora_id, bt.transportadora_id)))}</td>"
        f'<td style="text-align:center">{chip(bt.status.value.replace("_", " ").title(), _CLASSE_STATUS_TRANSP.get(bt.status.value, "neutro"))}</td>'
        "</tr>"
        for bt in bts
    ) or '<tr><td colspan="2" class="vazio">Nenhuma transportadora convidada.</td></tr>'

    linhas_ranking = "".join(
        "<tr>"
        f"<td>{i}</td>"
        f"<td>{escape(sc.nome)}</td>"
        f"<td>{numero(sc.score, 0)}</td>"
        f'<td style="text-align:center">{chip(sc.classificacao, _classe_score(sc.score))}</td>'
        f"<td>{numero(sc.proposta_media_rs_kg, 4)}</td>"
        f"<td>{numero(sc.prazo_medio, 0)} dias</td>"
        "</tr>"
        for i, sc in enumerate(scores, 1)
    ) or '<tr><td colspan="6" class="vazio">Nenhuma proposta recebida ainda.</td></tr>'

    bloco_economia = ""
    if eco:
        linhas_economia_grupo = "".join(
            "<tr>"
            f"<td>{escape(str(g['valor_grupo']))}</td>"
            f"<td>{numero(g['frete_atual_rs_kg'], 4)}</td>"
            f"<td>{numero(g['melhor_proposta_rs_kg'], 4)}</td>"
            f"<td>{moeda(g['economia'])}</td>"
            "</tr>"
            for g in eco.por_grupo
        ) or '<tr><td colspan="4" class="vazio">Sem dados por grupo.</td></tr>'
        bloco_economia = f"""
      <section class="bloco">
        <div class="titulo-secao"><h2>Potencial de Economia</h2></div>
        <div class="grade-kpi tabular">
          <div class="kpi"><span class="rotulo">Frete atual total</span><div class="valor">{moeda(eco.frete_atual_total)}</div></div>
          <div class="kpi"><span class="rotulo">Economia no período</span><div class="valor">{moeda(eco.economia_total_periodo)}</div><div class="legenda">{numero(eco.economia_pct, 1)}% de redução</div></div>
          <div class="kpi"><span class="rotulo">Ritmo mensal</span><div class="valor">{moeda(eco.ritmo_mensal)}</div></div>
          <div class="kpi"><span class="rotulo">Projeção anual</span><div class="valor">{moeda(eco.proj_anual)}</div></div>
        </div>
        <div class="tabela-wrap" style="margin-top:16px"><table class="tabular">
          <thead><tr><th>Grupo</th><th>R$/kg atual</th><th>Melhor proposta</th><th>Economia</th></tr></thead>
          <tbody>{linhas_economia_grupo}</tbody>
        </table></div>
      </section>
"""

    bloco_recomendacoes = _recomendacoes_bid(scores, eco, status_val)

    corpo_html = f"""
      <section class="bloco">
        <div class="titulo-secao"><h2>Escopo do BID</h2></div>
        <div class="tabela-wrap"><table class="tabular">
          <thead><tr><th>Agrupamento</th><th>Grupo</th><th>Embarques</th><th>Peso (kg)</th><th>Frete (R$)</th><th>R$/kg</th></tr></thead>
          <tbody>{linhas_escopo}</tbody>
        </table></div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Transportadoras Convidadas</h2></div>
        <div class="tabela-wrap"><table class="tabular">
          <thead><tr><th>Transportadora</th><th style="text-align:center">Status</th></tr></thead>
          <tbody>{linhas_transp}</tbody>
        </table></div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Ranking de Transportadoras</h2><span class="nota">Preço 40% + Prazo 30% + Cobertura 20% + Avaliação 10%</span></div>
        <div class="tabela-wrap"><table class="tabular">
          <thead><tr><th>#</th><th>Transportadora</th><th>Score</th><th style="text-align:center">Classificação</th><th>R$/kg médio</th><th>Prazo médio</th></tr></thead>
          <tbody>{linhas_ranking}</tbody>
        </table></div>
      </section>
{bloco_economia}
      <section class="bloco">
        <div class="titulo-secao"><h2>Recomendações</h2></div>
        {bloco_recomendacoes}
      </section>
"""

    return render_documento(
        titulo_pagina=f"Concorrência Logística — {escape(bid.nome)}",
        titulo_relatorio="Relatório Executivo de Concorrência Logística",
        meta_items=[
            f"<span>Empresa: <b>{escape(empresa_nome)}</b></span>",
            f"<span>BID: <b>{escape(bid.nome)}</b></span>",
            f"<span>Período: <b>{escape(str(bid.periodo_analise_inicio))} a {escape(str(bid.periodo_analise_fim))}</b></span>",
            f"<span>Gerado em: <b>{escape(gerado_em)}</b></span>",
        ],
        corpo_html=corpo_html,
        selo_saude=(rotulo_status, cor, fundo),
        resumo_executivo=_resumo_executivo_bid(escopos, scores, eco),
        rodape_nota="GD Conecta — Governança de Frete · Torre de Controle Logístico.",
    )
