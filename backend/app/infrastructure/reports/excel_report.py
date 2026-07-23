"""Estilos e helpers compartilhados dos relatórios em Excel (OpenPyXL) —
usados por benchmark_report.py e bid_report.py.

Paleta GD Conecta: Indigo #2D3561, Amber #C9A84C, Blue #0077A8, Ivory #F5F1EB.
"""
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

INDIGO = "2D3561"
AMBER = "C9A84C"
BLUE = "0077A8"
IVORY = "F5F1EB"

_header_fill = PatternFill("solid", fgColor=INDIGO)
_header_font = Font(color="FFFFFF", bold=True, size=11, name="Arial")
_title_font = Font(color=INDIGO, bold=True, size=16, name="Arial")
_sub_font = Font(color=BLUE, bold=True, size=12, name="Arial")
_thin = Side(style="thin", color="DDDDDD")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = _header_fill
        c.font = _header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border


def _autofit(ws):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 4, 45)
