"""Relatório do Fechamento de Custo de Frete (fechamento mensal), em HTML,
PDF e Excel — reaproveita exatamente os dados já montados por
``FechamentoMensalUseCase.obter``/``historico`` (Representatividade por
região e os totais nacionais), nunca recalculados em paralelo aqui.

O "Resumo geral" compara a competência atual com o mesmo mês do ano
anterior (mesma regra do card de KPIs da tela de Fechamento), redigido no
mesmo formato do e-mail que hoje é montado manualmente pelo analista
("Aumento de X%" / "Redução de X%").
"""
from html import escape
from io import BytesIO
from typing import Dict, List, Optional

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.infrastructure.reports.excel_report import _autofit, _style_header, _title_font
from app.infrastructure.reports.pdf_report import _br, _estilos, _tabela
from app.infrastructure.reports._html_theme import (
    OK, DANGER, SLATE, moeda, numero, pct, render_documento,
)

_ROTULO_MACRO = {
    "NORTE": "Norte", "NORDESTE": "Nordeste", "CENTRO_OESTE": "Centro-Oeste",
    "SUDESTE": "Sudeste", "SUL": "Sul",
}

_KPIS = [
    ("Faturamento", "faturamento_total", False),
    ("Devoluções", "devolucao", True),
    ("Frete contratado", "frete_contratado", True),
    ("Frete complementar", "frete_complementar", True),
]


def _rotulo_macro(m: str) -> str:
    return _ROTULO_MACRO.get(m, m)


def _competencia_ano_anterior(competencia: str) -> str:
    ano, mes = (int(p) for p in competencia.split("-"))
    return f"{ano - 1}-{mes:02d}"


def _rotulo_competencia(competencia: str) -> str:
    meses = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho",
             "agosto", "setembro", "outubro", "novembro", "dezembro"]
    ano, mes = (int(p) for p in competencia.split("-"))
    return f"{meses[mes - 1].capitalize()} {ano}"


def _buscar_ano_anterior(historico: List[Dict], competencia: str) -> Optional[Dict]:
    alvo = _competencia_ano_anterior(competencia)
    return next((h for h in historico if h["competencia"] == alvo), None)


def _comparativo(atual: Dict, anterior: Optional[Dict]) -> List[Dict]:
    """Monta os itens do Resumo Geral: rótulo, texto ("Aumento/Redução de
    X%") e se a variação é favorável (`bom`) — mesma regra do componente
    `Delta` da tela (inverte a leitura para Devoluções/Frete)."""
    itens = []
    for rotulo, campo, inverter in _KPIS:
        valor_atual = atual.get(campo) or 0.0
        valor_anterior = (anterior or {}).get(campo)
        if not valor_anterior:
            itens.append({"rotulo": rotulo, "texto": "sem base de comparação", "bom": None})
            continue
        delta = (valor_atual - valor_anterior) / valor_anterior * 100
        direcao = "Aumento" if delta >= 0 else "Redução"
        bom = (delta < 0) if inverter else (delta >= 0)
        itens.append({
            "rotulo": rotulo,
            "texto": f"{direcao} de {numero(abs(delta), 2)}%",
            "bom": bom,
        })
    return itens


# ============================ HTML ============================
def gerar_fechamento_html(
    empresa_nome: str,
    atual: Dict,
    historico: List[Dict],
    gerado_em: str,
) -> bytes:
    competencia = atual["competencia"]
    anterior = _buscar_ano_anterior(historico, competencia)
    comparativo = _comparativo(atual, anterior)

    cards_resumo = "".join(
        '<div class="kpi"><span class="rotulo">{rotulo}</span>'
        '<div class="valor" style="font-size:15px;color:{cor}">{texto}</div></div>'.format(
            rotulo=escape(item["rotulo"]),
            texto=escape(item["texto"]),
            cor=OK if item["bom"] else (DANGER if item["bom"] is False else SLATE),
        )
        for item in comparativo
    )

    linhas_regiao = "".join(
        "<tr>"
        f"<td>{escape(_rotulo_macro(r['macro_regiao']))}</td>"
        f"<td>{moeda(r['fat_real'])}</td>"
        f"<td>{moeda(r['frete_real'])}</td>"
        f"<td>{pct(r['pct_frete_fat_real'], 1)}</td>"
        f"<td>{moeda(r['fat_budget'])}</td>"
        f"<td>{moeda(r['frete_budget'])}</td>"
        f"<td>{pct(r['pct_frete_fat_budget'], 1)}</td>"
        f"<td>{pct(r['representatividade_real'], 1)}</td>"
        f"<td>{pct(r['representatividade_budget'], 1)}</td>"
        "</tr>"
        for r in atual["regional"]
    ) or '<tr><td colspan="9" class="vazio">Sem dados regionais.</td></tr>'

    tot = atual["regional_total"]
    linha_total = (
        "<tr style='font-weight:800'>"
        "<td>Total geral</td>"
        f"<td>{moeda(tot['fat_real'])}</td>"
        f"<td>{moeda(tot['frete_real'])}</td>"
        f"<td>{pct(tot['pct_frete_fat_real'], 1)}</td>"
        f"<td>{moeda(tot['fat_budget'])}</td>"
        f"<td>{moeda(tot['frete_budget'])}</td>"
        f"<td>{pct(tot['pct_frete_fat_budget'], 1)}</td>"
        "<td>100,00%</td><td>100,00%</td>"
        "</tr>"
    )

    corpo_html = f"""
      <section class="bloco">
        <div class="titulo-secao"><h2>Resumo geral</h2><span class="nota">vs. mesmo mês do ano anterior</span></div>
        <div class="grade-kpi tabular">
          {cards_resumo}
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Representatividade por região</h2></div>
        <p class="descricao-secao">Frete realizado e budget automáticos (CT-e e Metas Regionais); faturamento informado no fechamento.</p>
        <div class="tabela-wrap">
          <table class="tabular">
            <thead>
              <tr><th>Região</th><th>Vlr. Net Net</th><th>Frete líquido</th><th>% Frete/Fat.</th>
                  <th>Faturamento budget</th><th>Frete budget</th><th>% Frete/Fat. budget</th>
                  <th>Repr. real</th><th>Repr. budget</th></tr>
            </thead>
            <tbody>{linhas_regiao}{linha_total}</tbody>
          </table>
        </div>
      </section>
"""

    return render_documento(
        titulo_pagina=f"Fechamento de Custo de Frete — {escape(empresa_nome)}",
        titulo_relatorio=f"Fechamento de Custo de Frete — {_rotulo_competencia(competencia)}",
        meta_items=[
            f"<span>Empresa: <b>{escape(empresa_nome)}</b></span>",
            f"<span>Competência: <b>{_rotulo_competencia(competencia)}</b></span>",
            f"<span>Gerado em: <b>{escape(gerado_em)}</b></span>",
        ],
        corpo_html=corpo_html,
        rodape_nota="GD Conecta — Governança de Frete · Fechamento de Custo de Frete.",
    )


# ============================ PDF ============================
def gerar_fechamento_pdf(empresa_nome: str, atual: Dict, historico: List[Dict]) -> bytes:
    competencia = atual["competencia"]
    anterior = _buscar_ano_anterior(historico, competencia)
    comparativo = _comparativo(atual, anterior)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, topMargin=1.6 * cm, bottomMargin=1.6 * cm,
        leftMargin=1.6 * cm, rightMargin=1.6 * cm,
        title=f"Fechamento de Custo de Frete — {empresa_nome}",
    )
    s = _estilos()
    el = []
    el.append(Paragraph("Fechamento de Custo de Frete", s["GDTitulo"]))
    el.append(Paragraph(f"{empresa_nome} — {_rotulo_competencia(competencia)}", s["GDInfo"]))
    el.append(Spacer(1, 6))

    el.append(Paragraph("Resumo geral (vs. mesmo mês do ano anterior)", s["GDSub"]))
    el.append(_tabela(
        [["Indicador", "Variação"]] + [[item["rotulo"], item["texto"]] for item in comparativo],
        col_widths=[9 * cm, 7 * cm],
    ))

    el.append(Paragraph("Representatividade por região", s["GDSub"]))
    linhas = [["Região", "Fat. real", "Frete real", "% real", "Fat. budget", "Frete budget", "% budget", "Repr. real", "Repr. budget"]]
    for r in atual["regional"]:
        linhas.append([
            _rotulo_macro(r["macro_regiao"]), _br(r["fat_real"]), _br(r["frete_real"]),
            f"{_br(r['pct_frete_fat_real'], 1)}%", _br(r["fat_budget"]), _br(r["frete_budget"]),
            f"{_br(r['pct_frete_fat_budget'], 1)}%", f"{_br(r['representatividade_real'], 1)}%",
            f"{_br(r['representatividade_budget'], 1)}%",
        ])
    tot = atual["regional_total"]
    linhas.append([
        "Total geral", _br(tot["fat_real"]), _br(tot["frete_real"]), f"{_br(tot['pct_frete_fat_real'], 1)}%",
        _br(tot["fat_budget"]), _br(tot["frete_budget"]), f"{_br(tot['pct_frete_fat_budget'], 1)}%",
        "100,0%", "100,0%",
    ])
    el.append(_tabela(linhas))

    el.append(Spacer(1, 10))
    el.append(Paragraph("GD Conecta — Governança de Frete · Fechamento de Custo de Frete.", s["GDInfo"]))
    doc.build(el)
    return buf.getvalue()


# ============================ Excel ============================
def gerar_fechamento_excel(empresa_nome: str, atual: Dict, historico: List[Dict]) -> bytes:
    from openpyxl import Workbook

    competencia = atual["competencia"]
    anterior = _buscar_ano_anterior(historico, competencia)
    comparativo = _comparativo(atual, anterior)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.cell(row=1, column=1, value=f"Fechamento de Custo de Frete — {empresa_nome}").font = _title_font
    ws.cell(row=2, column=1, value=f"Competência: {_rotulo_competencia(competencia)}")
    ws.append([])
    ws.append(["Indicador", "Variação vs. mesmo mês ano anterior"])
    _style_header(ws, ws.max_row, 2)
    for item in comparativo:
        ws.append([item["rotulo"], item["texto"]])
    _autofit(ws)

    ws_reg = wb.create_sheet("Regional")
    ws_reg.append([
        "Região", "Faturamento real (R$)", "Frete real (R$)", "% Frete/Fat. real",
        "Faturamento budget (R$)", "Frete budget (R$)", "% Frete/Fat. budget",
        "Representatividade real", "Representatividade budget",
    ])
    _style_header(ws_reg, ws_reg.max_row, 9)
    for r in atual["regional"]:
        ws_reg.append([
            _rotulo_macro(r["macro_regiao"]), round(r["fat_real"], 2), round(r["frete_real"], 2),
            round(r["pct_frete_fat_real"], 2), round(r["fat_budget"], 2), round(r["frete_budget"], 2),
            round(r["pct_frete_fat_budget"], 2), round(r["representatividade_real"], 2),
            round(r["representatividade_budget"], 2),
        ])
    tot = atual["regional_total"]
    ws_reg.append([
        "Total geral", round(tot["fat_real"], 2), round(tot["frete_real"], 2),
        round(tot["pct_frete_fat_real"], 2), round(tot["fat_budget"], 2), round(tot["frete_budget"], 2),
        round(tot["pct_frete_fat_budget"], 2), 100.0, 100.0,
    ])
    _autofit(ws_reg)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
