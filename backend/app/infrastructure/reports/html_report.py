"""Geração do relatório de Diagnóstico Logístico em HTML e Excel.

Documento HTML autocontido (CSS inline, sem dependências externas) que reúne:
resumo executivo e selo de saúde geral, indicadores nacionais/regionais/
transportadoras (já existentes no relatório), indicadores por dimensão do DLG
(Filial/Rota/Transportadora/Cliente, top 10 mais críticos), outliers
detectados (apêndice, top 15) e as Recomendações consolidadas — reaproveita
exatamente os mesmos dados das telas Análise de Eficiência e Recomendações,
nunca recalcula nada em paralelo. O Excel reaproveita os mesmos dados em
planilhas separadas, com as listas completas (sem os recortes top 10/15 do HTML).

Todo texto de origem do usuário (nomes de cliente/transportadora vindos de
CT-e importado) passa por ``html.escape`` — o relatório é aberto direto no
navegador, então texto não escapado seria um vetor de XSS armazenado.
"""
from html import escape
from io import BytesIO
from typing import Dict, List, Optional

from app.application.dtos import Diagnostico
from app.infrastructure.reports._html_theme import (
    DANGER, DANGER_BG, OK, OK_BG, WARN, WARN_BG,
    chip, moeda, numero, pct, render_documento,
)
from app.infrastructure.reports.excel_report import _autofit, _style_header, _title_font

_ROTULO_DIMENSAO = {
    "FILIAL": "Filial",
    "ROTA": "Rota",
    "TRANSPORTADORA": "Transportadora",
    "CLIENTE_FINAL": "Cliente",
}
_ROTULO_CLASSIFICACAO = {
    "CRITICO": "Crítico", "ATENCAO": "Atenção",
    "EFICIENTE": "Eficiente", "SEM_REF": "Sem ref.",
}
_CLASSE_CLASSIFICACAO = {
    "CRITICO": "critico", "ATENCAO": "atencao",
    "EFICIENTE": "eficiente", "SEM_REF": "neutro",
}
_ROTULO_PRIORIDADE = {"CRITICA": "Crítica", "ALTA": "Alta", "MEDIA": "Média", "BAIXA": "Baixa"}
_CLASSE_PRIORIDADE = {"CRITICA": "critico", "ALTA": "critico", "MEDIA": "atencao", "BAIXA": "neutro"}
_ROTULO_TIPO_OUTLIER = {
    "RS_KG_2DP": "R$/kg > média+2σ",
    "PCT_FRETE": "% Frete > limite",
    "VAR_MENSAL": "Variação mensal > 30%",
}


def _top_criticos(linhas: List[dict], n: int = 10) -> List[dict]:
    candidatos = [r for r in linhas if r.get("classificacao") in ("CRITICO", "ATENCAO")]
    return sorted(candidatos, key=lambda r: r.get("desvio_pct") or 0, reverse=True)[:n]


def _distribuicao(linhas: List[dict]) -> Dict[str, int]:
    d = {"CRITICO": 0, "ATENCAO": 0, "EFICIENTE": 0, "SEM_REF": 0}
    for r in linhas:
        d[r.get("classificacao", "SEM_REF")] = d.get(r.get("classificacao", "SEM_REF"), 0) + 1
    return d


def _saude_geral(dlg_por_dimensao: Dict[str, List[dict]]) -> tuple:
    """Selo de saúde geral do diagnóstico, a partir da distribuição de
    classificações já calculada por dimensão (sem recalcular nada). Limiares:
    >=15% de registros críticos = Crítico; algum crítico ou >=25% em atenção
    = Atenção; caso contrário, Saudável."""
    total = {"CRITICO": 0, "ATENCAO": 0, "EFICIENTE": 0, "SEM_REF": 0}
    for linhas in dlg_por_dimensao.values():
        for k, v in _distribuicao(linhas).items():
            total[k] += v
    n = sum(total.values()) or 1
    pct_critico = total["CRITICO"] / n
    pct_atencao = total["ATENCAO"] / n
    if pct_critico >= 0.15:
        return ("Crítico", DANGER, DANGER_BG)
    if total["CRITICO"] > 0 or pct_atencao >= 0.25:
        return ("Atenção", WARN, WARN_BG)
    return ("Saudável", OK, OK_BG)


def _resumo_executivo(diag: Diagnostico, outliers: List, recomendacoes: List) -> str:
    n = diag.nacional
    partes = [f"{n.qtd_ctes} CT-e analisados no período"]
    if n.qtd_ctes and outliers:
        partes.append(f"{pct(len(outliers) / n.qtd_ctes * 100, 1)} fora do padrão estatístico")
    economia_total = sum((r.economia_estimada or 0) for r in recomendacoes)
    if economia_total:
        partes.append(f"economia potencial estimada em {moeda(economia_total)}")
    return escape(" — ".join(partes)) + "."


def _linha_dimensao_html(linhas_dim: List[dict]) -> str:
    if not linhas_dim:
        return '<p class="vazio">Sem registros críticos ou em atenção nesta dimensão no período.</p>'
    linhas_html = []
    for r in linhas_dim:
        clf = r.get("classificacao", "SEM_REF")
        linhas_html.append(
            "<tr>"
            f'<td>{escape(str(r.get("chave_label") or "—"))}</td>'
            f'<td>{r.get("qtd_ctes", 0)}</td>'
            f'<td>{moeda(r.get("rs_kg"))}</td>'
            f'<td class="{"desvio-pos" if (r.get("desvio_pct") or 0) > 0 else "desvio-neg"}">'
            f'{"+" if (r.get("desvio_pct") or 0) > 0 else ""}{numero(r.get("desvio_pct"), 1)}%</td>'
            f'<td style="text-align:center">{chip(_ROTULO_CLASSIFICACAO.get(clf, clf), _CLASSE_CLASSIFICACAO.get(clf, "neutro"))}</td>'
            "</tr>"
        )
    return (
        '<div class="tabela-wrap"><table class="tabular"><thead><tr>'
        "<th>Entidade</th><th>CT-e</th><th>R$/kg</th><th>Desvio</th>"
        '<th style="text-align:center">Classif.</th>'
        f"</tr></thead><tbody>{''.join(linhas_html)}</tbody></table></div>"
    )


def gerar_relatorio_html(
    diag: Diagnostico,
    dlg_por_dimensao: Dict[str, List[dict]],
    outliers: List,
    recomendacoes: List,
    gerado_em: str,
) -> bytes:
    n = diag.nacional
    periodo = "Todo o período"
    if diag.periodo_inicio or diag.periodo_fim:
        periodo = f"{diag.periodo_inicio or '...'} a {diag.periodo_fim or '...'}"

    # ---------- Indicadores por região ----------
    linhas_regiao = "".join(
        "<tr>"
        f"<td>{escape(r.macro_regiao)}</td><td>{r.qtd_ctes}</td>"
        f"<td>{moeda(r.frete_rs_kg)}</td><td>{pct(r.frete_pct)}</td>"
        f"<td>{moeda(r.frete_total)}</td>"
        "</tr>"
        for r in diag.regionais
    ) or '<tr><td colspan="5" class="vazio">Sem dados regionais.</td></tr>'

    # ---------- Indicadores por dimensão (Protótipo A) ----------
    secoes_dimensao = []
    for dim in ("CLIENTE_FINAL", "ROTA", "TRANSPORTADORA", "FILIAL"):
        linhas_dim = dlg_por_dimensao.get(dim, [])
        dist = _distribuicao(linhas_dim)
        top = _top_criticos(linhas_dim, n=10)
        secoes_dimensao.append(
            f'<div class="subgrupo"><span class="rotulo-dim">{escape(_ROTULO_DIMENSAO[dim])} '
            f'&middot; {len(linhas_dim)} registro(s) no período</span>'
            f'<div class="distribuicao">'
            f'<span class="item"><span class="ponto" style="background:{DANGER}"></span><b>{dist["CRITICO"]}</b>&nbsp;crítico</span>'
            f'<span class="item"><span class="ponto" style="background:{WARN}"></span><b>{dist["ATENCAO"]}</b>&nbsp;atenção</span>'
            f'<span class="item"><span class="ponto" style="background:{OK}"></span><b>{dist["EFICIENTE"]}</b>&nbsp;eficiente</span>'
            f"</div>{_linha_dimensao_html(top)}</div>"
        )
    bloco_dimensao = "".join(secoes_dimensao)

    # ---------- Outliers (apêndice) ----------
    top_outliers = sorted(outliers, key=lambda o: o.desvio_pct, reverse=True)[:15]
    if top_outliers:
        linhas_out = "".join(
            "<tr>"
            f"<td>{escape(o.periodo_ref)}</td>"
            f'<td>{chip(_ROTULO_TIPO_OUTLIER.get(o.tipo_outlier, o.tipo_outlier), "neutro")}</td>'
            f"<td>{numero(o.valor_real, 2)}</td><td>{numero(o.valor_limite, 2)}</td>"
            f'<td class="desvio-pos">+{numero(o.desvio_pct, 1)}%</td><td>{o.cte_id}</td>'
            "</tr>"
            for o in top_outliers
        )
        bloco_outliers = (
            '<div class="tabela-wrap"><table class="tabular"><thead><tr>'
            "<th>Período</th><th>Tipo</th><th>Valor real</th><th>Limite</th>"
            "<th>Desvio</th><th>CT-e nº</th></tr></thead>"
            f"<tbody>{linhas_out}</tbody></table></div>"
        )
    else:
        bloco_outliers = '<p class="vazio">Nenhum outlier detectado no período.</p>'

    # ---------- Recomendações ----------
    if recomendacoes:
        blocos_reco = []
        for r in recomendacoes:
            prioridade = r.prioridade
            acao = (r.dados or {}).get("acao_recomendada") if r.dados else None
            valores = []
            if r.valor_encontrado:
                valores.append(f'<span class="par">Encontrado: <b>{escape(r.valor_encontrado)}</b></span>')
            if r.valor_recomendado:
                valores.append(f'<span class="par">Referência: <b>{escape(r.valor_recomendado)}</b></span>')
            if r.economia_estimada:
                valores.append(f'<span class="par economia">Economia potencial: {moeda(r.economia_estimada)}</span>')
            blocos_reco.append(
                f'<div class="recomendacao {_CLASSE_PRIORIDADE.get(prioridade, "neutro")}">'
                '<div class="linha-topo">'
                f"<h3>{escape(r.titulo)}</h3>"
                f'{chip(_ROTULO_PRIORIDADE.get(prioridade, prioridade), _CLASSE_PRIORIDADE.get(prioridade, "neutro"))}'
                "</div>"
                f"<p>{escape(r.descricao)}</p>"
                + (f'<div class="acao"><b>O que fazer:</b> {escape(acao)}</div>' if acao else "")
                + (f'<div class="valores tabular">{"".join(valores)}</div>' if valores else "")
                + "</div>"
            )
        bloco_recomendacoes = "".join(blocos_reco)
    else:
        bloco_recomendacoes = (
            '<p class="vazio">Nenhuma recomendação consolidada ainda — '
            "gere-as na tela Recomendações antes de exportar o relatório.</p>"
        )

    corpo_html = f"""
      <section class="bloco">
        <div class="titulo-secao"><h2>Indicadores Nacionais</h2></div>
        <div class="grade-kpi tabular">
          <div class="kpi"><span class="rotulo">CT-e analisados</span><div class="valor">{n.qtd_ctes}</div></div>
          <div class="kpi"><span class="rotulo">Frete total</span><div class="valor">{moeda(n.valor_total_frete)}</div></div>
          <div class="kpi"><span class="rotulo">Custo por kg</span><div class="valor">{moeda(n.frete_rs_kg)}</div>
            {f'<div class="legenda {"desvio-neg" if (n.desvio_rs_kg or 0) <= 0 else "desvio-pos"}">{numero(n.desvio_rs_kg, 1)}% vs. meta</div>' if n.desvio_rs_kg is not None else '<div class="legenda">sem meta definida</div>'}
          </div>
          <div class="kpi"><span class="rotulo">% Frete / mercadoria</span><div class="valor">{pct(n.frete_pct)}</div>
            {f'<div class="legenda">meta: {pct(n.meta_pct)}</div>' if n.meta_pct is not None else '<div class="legenda">sem meta definida</div>'}
          </div>
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Indicadores por Região</h2></div>
        <div class="tabela-wrap">
          <table class="tabular">
            <thead><tr><th>Região</th><th>CT-e</th><th>R$/kg</th><th>% Frete</th><th>Frete total</th></tr></thead>
            <tbody>{linhas_regiao}</tbody>
          </table>
        </div>
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Indicadores por Dimensão</h2><span class="nota">top 10 mais críticos por dimensão</span></div>
        <p class="descricao-secao">Mesma análise da tela Diagnóstico Logístico → Análise de Eficiência, condensada aos casos que mais pesam no custo.</p>
        {bloco_dimensao}
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Outliers Detectados</h2><span class="nota">top 15 por desvio</span></div>
        <p class="descricao-secao">CT-e individuais fora do padrão estatístico da empresa — mesmo critério da aba Outliers da Análise de Eficiência.</p>
        {bloco_outliers}
      </section>

      <section class="bloco">
        <div class="titulo-secao"><h2>Recomendações</h2></div>
        <p class="descricao-secao">Oportunidades consolidadas a partir dos indicadores acima — gerencie o status de cada uma na tela Recomendações.</p>
        {bloco_recomendacoes}
      </section>
"""

    return render_documento(
        titulo_pagina=f"Diagnóstico Logístico — {escape(diag.empresa_nome)}",
        titulo_relatorio="Relatório de Diagnóstico Logístico",
        meta_items=[
            f"<span>Empresa: <b>{escape(diag.empresa_nome)}</b></span>",
            f"<span>Período: <b>{escape(periodo)}</b></span>",
            f"<span>Gerado em: <b>{escape(gerado_em)}</b></span>",
        ],
        corpo_html=corpo_html,
        selo_saude=_saude_geral(dlg_por_dimensao),
        resumo_executivo=_resumo_executivo(diag, outliers, recomendacoes),
        rodape_nota="GD Diagnóstico Logístico — relatório gerado automaticamente a partir dos indicadores do Diagnóstico Logístico.",
    )


def gerar_relatorio_excel(
    diag: Diagnostico,
    dlg_por_dimensao: Dict[str, List[dict]],
    outliers: List,
    recomendacoes: List,
) -> bytes:
    from openpyxl import Workbook

    n = diag.nacional
    periodo = "Todo o período"
    if diag.periodo_inicio or diag.periodo_fim:
        periodo = f"{diag.periodo_inicio or '...'} a {diag.periodo_fim or '...'}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Nacional"
    ws.cell(row=1, column=1, value=f"Diagnóstico Logístico — {diag.empresa_nome}").font = _title_font
    ws.cell(row=2, column=1, value=f"Período: {periodo}")
    ws.append([])
    ws.append(["Indicador", "Valor"])
    _style_header(ws, ws.max_row, 2)
    ws.append(["CT-e analisados", n.qtd_ctes])
    ws.append(["Frete total (R$)", round(n.valor_total_frete, 2)])
    ws.append(["Custo por kg (R$)", round(n.frete_rs_kg, 4)])
    ws.append(["% Frete / mercadoria", round(n.frete_pct, 2)])
    if n.meta_pct is not None:
        ws.append(["Meta % Frete", round(n.meta_pct, 2)])
    if n.desvio_rs_kg is not None:
        ws.append(["Desvio vs. meta (%)", round(n.desvio_rs_kg, 1)])
    _autofit(ws)

    ws_reg = wb.create_sheet("Regional")
    ws_reg.append(["Região", "CT-e", "R$/kg", "% Frete", "Frete total (R$)"])
    _style_header(ws_reg, ws_reg.max_row, 5)
    for r in diag.regionais:
        ws_reg.append([r.macro_regiao, r.qtd_ctes, round(r.frete_rs_kg, 4), round(r.frete_pct, 2), round(r.frete_total, 2)])
    _autofit(ws_reg)

    for dim in ("CLIENTE_FINAL", "ROTA", "TRANSPORTADORA", "FILIAL"):
        linhas = dlg_por_dimensao.get(dim, [])
        ws_dim = wb.create_sheet(_ROTULO_DIMENSAO[dim])
        ws_dim.append(["Entidade", "CT-e", "R$/kg", "Desvio %", "Classificação"])
        _style_header(ws_dim, ws_dim.max_row, 5)
        for r in linhas:
            clf = r.get("classificacao", "SEM_REF")
            ws_dim.append([
                str(r.get("chave_label") or "—"), r.get("qtd_ctes", 0),
                round(r.get("rs_kg") or 0, 4), round(r.get("desvio_pct") or 0, 1),
                _ROTULO_CLASSIFICACAO.get(clf, clf),
            ])
        _autofit(ws_dim)

    ws_out = wb.create_sheet("Outliers")
    ws_out.append(["Período", "Tipo", "Valor real", "Limite", "Desvio %", "CT-e nº"])
    _style_header(ws_out, ws_out.max_row, 6)
    for o in sorted(outliers, key=lambda o: o.desvio_pct, reverse=True):
        ws_out.append([
            o.periodo_ref, _ROTULO_TIPO_OUTLIER.get(o.tipo_outlier, o.tipo_outlier),
            round(o.valor_real, 2), round(o.valor_limite, 2), round(o.desvio_pct, 1), o.cte_id,
        ])
    _autofit(ws_out)

    ws_rec = wb.create_sheet("Recomendações")
    ws_rec.append(["Título", "Prioridade", "Descrição", "Encontrado", "Referência", "Economia estimada (R$)"])
    _style_header(ws_rec, ws_rec.max_row, 6)
    for r in recomendacoes:
        ws_rec.append([
            r.titulo, _ROTULO_PRIORIDADE.get(r.prioridade, r.prioridade), r.descricao,
            r.valor_encontrado or "", r.valor_recomendado or "",
            round(r.economia_estimada, 2) if r.economia_estimada else None,
        ])
    _autofit(ws_rec)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
