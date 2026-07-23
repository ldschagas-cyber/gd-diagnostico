"""Routers de configuração do Benchmark OD (V2.1).

- /empresas/{id}/hubs             → catálogo de hubs, próprio de cada empresa
- /empresas/{id}/clusters         → mapa UF/município → hub, por empresa
- /empresas/{id}/corredores       → referência de mercado por corredor, por empresa
"""
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from app.domain.entities import BenchmarkCorredor, ClusterCliente, HubLogistico
from app.presentation.api.dependencies import (
    bloquear_visualizador,
    get_benchmark_corredor_repo,
    get_cluster_repo,
    get_empresa_repo,
    get_hub_repo,
    verificar_acesso_empresa,
)
from app.presentation.schemas import (
    BenchmarkCorredorIn,
    BenchmarkCorredorOut,
    ClusterClienteIn,
    ClusterClienteOut,
    HubLogisticoIn,
    HubLogisticoOut,
)

# ───────────────────────── Hubs (catálogo por empresa) ─────────────────────────
hubs_router = APIRouter(prefix="/empresas", tags=["Benchmark OD — Hubs"])


@hubs_router.get("/{empresa_id}/hubs", response_model=list[HubLogisticoOut])
def listar_hubs(
    empresa_id: int,
    apenas_ativos: bool = False,
    _=Depends(verificar_acesso_empresa),
    repo=Depends(get_hub_repo),
):
    return repo.list(empresa_id, apenas_ativos=apenas_ativos)


@hubs_router.post("/{empresa_id}/hubs", response_model=HubLogisticoOut, status_code=201)
def criar_hub(
    empresa_id: int,
    payload: HubLogisticoIn,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_hub_repo),
):
    codigo = payload.codigo.strip().upper().replace(" ", "_")
    if repo.get_by_codigo(empresa_id, codigo):
        raise HTTPException(409, f"Já existe um hub com o código {codigo} nesta empresa.")
    dados = payload.model_dump()
    dados["codigo"] = codigo
    return repo.create(HubLogistico(empresa_id=empresa_id, **dados))


@hubs_router.put("/{empresa_id}/hubs/{hub_id}", response_model=HubLogisticoOut)
def atualizar_hub(
    empresa_id: int,
    hub_id: int,
    payload: HubLogisticoIn,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_hub_repo),
):
    atual = repo.get(hub_id)
    if not atual or atual.empresa_id != empresa_id:
        raise HTTPException(404, "Hub não encontrado para esta empresa.")
    dados = payload.model_dump()
    dados["codigo"] = payload.codigo.strip().upper().replace(" ", "_")
    atualizado = repo.update(hub_id, HubLogistico(empresa_id=empresa_id, **dados))
    return atualizado


@hubs_router.delete("/{empresa_id}/hubs/{hub_id}", status_code=204)
def remover_hub(
    empresa_id: int,
    hub_id: int,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_hub_repo),
):
    atual = repo.get(hub_id)
    if not atual or atual.empresa_id != empresa_id:
        raise HTTPException(404, "Hub não encontrado para esta empresa.")
    repo.delete(hub_id)


# ─────────────────── Clusters por empresa-cliente ───────────────────
clusters_router = APIRouter(prefix="/empresas", tags=["Benchmark OD — Clusters"])


@clusters_router.get("/{empresa_id}/clusters", response_model=list[ClusterClienteOut])
def listar_clusters(
    empresa_id: int,
    _=Depends(verificar_acesso_empresa),
    repo=Depends(get_cluster_repo),
    empresa_repo=Depends(get_empresa_repo),
):
    if not empresa_repo.get(empresa_id):
        raise HTTPException(404, "Empresa não encontrada.")
    return repo.list_by_empresa(empresa_id)


@clusters_router.post("/{empresa_id}/clusters", response_model=ClusterClienteOut, status_code=201)
def criar_cluster(
    empresa_id: int,
    payload: ClusterClienteIn,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_cluster_repo),
    hub_repo=Depends(get_hub_repo),
    empresa_repo=Depends(get_empresa_repo),
):
    if not empresa_repo.get(empresa_id):
        raise HTTPException(404, "Empresa não encontrada.")
    hub = hub_repo.get(payload.hub_id)
    if not hub or hub.empresa_id != empresa_id:
        raise HTTPException(422, "Hub inválido.")
    cluster = ClusterCliente(empresa_id=empresa_id, **payload.model_dump())
    return repo.create(cluster)


@clusters_router.get("/{empresa_id}/clusters/modelo")
def baixar_modelo_clusters(
    empresa_id: int,
    _=Depends(verificar_acesso_empresa),
):
    """Baixa a planilha-modelo (.xlsx) para importação de clusters logísticos.
    Colunas: uf, municipio (opcional), hub_codigo."""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clusters"
    ws.append(["uf", "municipio", "hub_codigo"])
    azul = PatternFill("solid", fgColor="2D3561")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
    ws.append(["SP", "São Paulo", "SUDESTE_HUB"])
    ws.append(["SP", "", "SUDESTE_HUB"])
    ws.append(["PE", "Recife", "NORDESTE_HUB"])
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22

    inst = wb.create_sheet("Instruções")
    for linha in [
        ["Como preencher"],
        [""],
        ["uf", "Sigla da UF com 2 letras, ex.: SP, PE (obrigatório)."],
        ["municipio", "Nome do município (opcional). Vazio = regra para a UF inteira."],
        ["hub_codigo", "Código de um Hub já cadastrado desta empresa, ex.: SUDESTE_HUB (obrigatório)."],
        [""],
        ["Clusters já cadastrados (mesma UF + município) são ignorados."],
    ]:
        inst.append(linha)
    inst["A1"].font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 18
    inst.column_dimensions["B"].width = 70

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_clusters.xlsx"},
    )


@clusters_router.post("/{empresa_id}/clusters/importar", response_model=dict)
async def importar_clusters_excel(
    empresa_id: int,
    arquivo: UploadFile = File(...),
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_cluster_repo),
    hub_repo=Depends(get_hub_repo),
    empresa_repo=Depends(get_empresa_repo),
):
    """Importa clusters via Excel (.xlsx). Colunas: uf, municipio (opcional),
    hub_codigo. Valida duplicidade (empresa + UF + município) e hub existente."""
    import openpyxl
    from io import BytesIO
    from app.infrastructure.parsers.macro_regiao import macro_por_uf

    if not empresa_repo.get(empresa_id):
        raise HTTPException(404, "Empresa não encontrada.")

    conteudo = await arquivo.read()
    if not conteudo.startswith(b"PK"):
        raise HTTPException(400, "Arquivo não é um Excel (.xlsx) válido.")
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    if "uf" not in col or "hub_codigo" not in col:
        raise HTTPException(400, "Colunas obrigatórias ausentes: uf, hub_codigo.")

    # dedup contra o que já existe + dentro do próprio arquivo
    existentes = {(c.uf.upper(), (c.municipio or "").strip().lower())
                  for c in repo.list_by_empresa(empresa_id)}
    vistos: set[tuple] = set()
    importados = 0
    ignorados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        uf = str(row[col["uf"]] or "").strip().upper()
        municipio = str(row[col["municipio"]] or "").strip() if "municipio" in col else ""
        cod = str(row[col["hub_codigo"]] or "").strip().upper()
        if macro_por_uf(uf) is None or not cod:
            ignorados += 1  # UF inválida ou hub vazio
            continue
        hub = hub_repo.get_by_codigo(empresa_id, cod)
        if not hub:
            ignorados += 1  # hub inexistente (ou de outra empresa)
            continue
        chave = (uf, municipio.lower())
        if chave in existentes or chave in vistos:
            ignorados += 1  # duplicado (banco ou planilha)
            continue
        vistos.add(chave)
        repo.create(ClusterCliente(
            empresa_id=empresa_id, uf=uf, municipio=municipio, hub_id=hub.id,
        ))
        importados += 1

    return {"importados": importados, "ignorados": ignorados}


@clusters_router.put("/{empresa_id}/clusters/{cluster_id}", response_model=ClusterClienteOut)
def atualizar_cluster(
    empresa_id: int,
    cluster_id: int,
    payload: ClusterClienteIn,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_cluster_repo),
    hub_repo=Depends(get_hub_repo),
):
    atual = repo.get(cluster_id)
    if not atual or atual.empresa_id != empresa_id:
        raise HTTPException(404, "Cluster não encontrado para esta empresa.")
    hub = hub_repo.get(payload.hub_id)
    if not hub or hub.empresa_id != empresa_id:
        raise HTTPException(422, "Hub inválido.")
    cluster = ClusterCliente(empresa_id=empresa_id, **payload.model_dump())
    return repo.update(cluster_id, cluster)


@clusters_router.delete("/{empresa_id}/clusters/{cluster_id}", status_code=204)
def remover_cluster(
    empresa_id: int,
    cluster_id: int,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_cluster_repo),
):
    atual = repo.get(cluster_id)
    if not atual or atual.empresa_id != empresa_id:
        raise HTTPException(404, "Cluster não encontrado para esta empresa.")
    repo.delete(cluster_id)


# ─────────────── Referências de corredor (por empresa) ───────────────
corredor_router = APIRouter(prefix="/empresas", tags=["Benchmark OD — Referências"])


@corredor_router.get("/{empresa_id}/corredores", response_model=list[BenchmarkCorredorOut])
def listar_corredores(
    empresa_id: int,
    _=Depends(verificar_acesso_empresa),
    repo=Depends(get_benchmark_corredor_repo),
):
    return repo.list(empresa_id)


@corredor_router.put("/{empresa_id}/corredores", response_model=BenchmarkCorredorOut)
def upsert_corredor(
    empresa_id: int,
    payload: BenchmarkCorredorIn,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_benchmark_corredor_repo),
):
    """Cria ou atualiza a referência de um corredor (upsert por par de hubs)."""
    return repo.upsert(BenchmarkCorredor(empresa_id=empresa_id, **payload.model_dump()))


@corredor_router.get("/{empresa_id}/corredores/modelo")
def baixar_modelo_corredores(
    empresa_id: int,
    _=Depends(verificar_acesso_empresa),
):
    """Baixa a planilha-modelo (.xlsx) para importação de referências de
    corredor. Colunas: hub_origem_codigo, hub_destino_codigo (obrigatórias),
    frete_kg_min/medio/max, frete_pct_min/medio/max, volume_referencia,
    dispersao_kg, observacoes (opcionais, default 0/vazio)."""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corredores"
    colunas = [
        "hub_origem_codigo", "hub_destino_codigo",
        "frete_kg_min", "frete_kg_medio", "frete_kg_max",
        "frete_pct_min", "frete_pct_medio", "frete_pct_max",
        "volume_referencia", "dispersao_kg", "prazo_dias_medio", "observacoes",
    ]
    ws.append(colunas)
    azul = PatternFill("solid", fgColor="2D3561")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
    ws.append(["SUDESTE_HUB", "NORDESTE_HUB", 0.8, 1.2, 1.8, 3.0, 6.0, 12.0, 5000, 0.15, 3, ""])
    ws.append(["SUDESTE_HUB", "SUL_HUB", 0.6, 0.9, 1.3, 2.0, 4.5, 9.0, 3000, 0.10, 2, "Rota consolidada"])
    for i, largura in enumerate([18, 18, 11, 11, 11, 11, 11, 11, 16, 12, 14, 26], start=1):
        ws.column_dimensions[chr(64 + i)].width = largura

    inst = wb.create_sheet("Instruções")
    for linha in [
        ["Como preencher"],
        [""],
        ["hub_origem_codigo", "Código de um Hub já cadastrado desta empresa (obrigatório)."],
        ["hub_destino_codigo", "Código de um Hub já cadastrado desta empresa (obrigatório)."],
        ["frete_kg_min/medio/max", "Faixa de R$/kg de referência do corredor (opcional, default 0)."],
        ["frete_pct_min/medio/max", "Faixa de % Frete/mercadoria de referência (opcional, default 0)."],
        ["volume_referencia", "Peso de mercado do fluxo, usado como ponderação (opcional, default 0)."],
        ["dispersao_kg", "Tolerância de dispersão do corredor, entre 0 e 1 (opcional, default 0)."],
        ["prazo_dias_medio", "Prazo de entrega de referência do corredor, em dias (opcional, default 0 = não cadastrado)."],
        ["observacoes", "Texto livre (opcional)."],
        [""],
        ["Um corredor já cadastrado (mesmo par de hubs origem→destino) é ATUALIZADO, não duplicado."],
    ]:
        inst.append(linha)
    inst["A1"].font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 22
    inst.column_dimensions["B"].width = 70

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_corredores.xlsx"},
    )


@corredor_router.post("/{empresa_id}/corredores/importar", response_model=dict)
async def importar_corredores_excel(
    empresa_id: int,
    arquivo: UploadFile = File(...),
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_benchmark_corredor_repo),
    hub_repo=Depends(get_hub_repo),
    empresa_repo=Depends(get_empresa_repo),
):
    """Importa referências de corredor via Excel (.xlsx). Colunas obrigatórias:
    hub_origem_codigo, hub_destino_codigo. Demais colunas numéricas são
    opcionais (default 0). Um par de hubs já cadastrado é ATUALIZADO (upsert),
    não duplicado nem ignorado — diferente do import de clusters."""
    import openpyxl
    from io import BytesIO

    if not empresa_repo.get(empresa_id):
        raise HTTPException(404, "Empresa não encontrada.")

    conteudo = await arquivo.read()
    if not conteudo.startswith(b"PK"):
        raise HTTPException(400, "Arquivo não é um Excel (.xlsx) válido.")
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    if "hub_origem_codigo" not in col or "hub_destino_codigo" not in col:
        raise HTTPException(400, "Colunas obrigatórias ausentes: hub_origem_codigo, hub_destino_codigo.")

    def _num(row, chave):
        if chave not in col:
            return 0.0
        v = row[col[chave]]
        try:
            return float(v) if v is not None else 0.0
        except (TypeError, ValueError):
            return 0.0

    existentes = {(c.hub_origem_codigo, c.hub_destino_codigo) for c in repo.list(empresa_id)}
    importados = 0
    atualizados = 0
    ignorados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        origem = str(row[col["hub_origem_codigo"]] or "").strip().upper()
        destino = str(row[col["hub_destino_codigo"]] or "").strip().upper()
        if not origem or not destino:
            ignorados += 1
            continue
        if not hub_repo.get_by_codigo(empresa_id, origem) or not hub_repo.get_by_codigo(empresa_id, destino):
            ignorados += 1  # hub de origem/destino inexistente (ou de outra empresa)
            continue
        chave = (origem, destino)
        observacoes = str(row[col["observacoes"]] or "").strip() if "observacoes" in col else ""
        repo.upsert(BenchmarkCorredor(
            empresa_id=empresa_id,
            hub_origem_codigo=origem, hub_destino_codigo=destino,
            frete_kg_min=_num(row, "frete_kg_min"),
            frete_kg_medio=_num(row, "frete_kg_medio"),
            frete_kg_max=_num(row, "frete_kg_max"),
            frete_pct_min=_num(row, "frete_pct_min"),
            frete_pct_medio=_num(row, "frete_pct_medio"),
            frete_pct_max=_num(row, "frete_pct_max"),
            volume_referencia=_num(row, "volume_referencia"),
            dispersao_kg=_num(row, "dispersao_kg"),
            prazo_dias_medio=_num(row, "prazo_dias_medio"),
            observacoes=observacoes,
        ))
        if chave in existentes:
            atualizados += 1
        else:
            importados += 1
            existentes.add(chave)

    return {"importados": importados, "atualizados": atualizados, "ignorados": ignorados}


@corredor_router.delete("/{empresa_id}/corredores/{corredor_id}", status_code=204)
def remover_corredor(
    empresa_id: int,
    corredor_id: int,
    _ac=Depends(verificar_acesso_empresa),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_benchmark_corredor_repo),
):
    atual = repo.get(corredor_id)
    if not atual or atual.empresa_id != empresa_id:
        raise HTTPException(404, "Referência de corredor não encontrada.")
    repo.delete(corredor_id)
