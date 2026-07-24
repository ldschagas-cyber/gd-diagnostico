"""Benchmark V2 — endpoints (Freight Market Intelligence).

Expõe:
- Geração e leitura do benchmark OBSERVADO (percentis reais por OD).
- Comparação observado vs mercado.
- Leitura e EDIÇÃO da matriz de MERCADO (source of truth — Matriz OD).
- Resolver de referência (fallback cliente → mercado).
"""
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.presentation.api.dependencies import (
    bloquear_visualizador,
    get_current_superuser,
    get_current_user,
    verificar_acesso_empresa,
)
from app.application.use_cases.benchmark_v2 import BenchmarkV2UseCase
from app.application.use_cases.benchmark_observado import BenchmarkObservadoUseCase
from app.application.use_cases.indicadores_regionais import IndicadoresRegionaisUseCase
from app.infrastructure.database.models import BenchmarkMercadoModel

router = APIRouter(prefix="/benchmark-v2", tags=["Benchmark V2 — Market Intelligence"])


def _obs_dict(o) -> dict:
    return {
        "origem_regiao": o.origem_regiao, "destino_regiao": o.destino_regiao,
        "qtd_embarques": o.qtd_embarques,
        "rs_kg_p10": o.rs_kg_p10, "rs_kg_p25": o.rs_kg_p25, "rs_kg_p50": o.rs_kg_p50,
        "rs_kg_p75": o.rs_kg_p75, "rs_kg_p90": o.rs_kg_p90,
        "media": o.media, "desvio_padrao": o.desvio_padrao, "periodo_ref": o.periodo_ref,
    }


def _merc_dict(m) -> dict:
    return {
        "id": m.id, "origem_regiao": m.origem_regiao, "destino_regiao": m.destino_regiao,
        "modal": m.modal, "tipo_operacao": m.tipo_operacao, "setor": m.setor,
        "faixa_peso_min": m.faixa_peso_min, "faixa_peso_max": m.faixa_peso_max,
        "rs_kg_p10": m.rs_kg_p10, "rs_kg_p25": m.rs_kg_p25, "rs_kg_p50": m.rs_kg_p50,
        "rs_kg_p75": m.rs_kg_p75, "rs_kg_p90": m.rs_kg_p90, "fonte": m.fonte,
        "frete_pct_min": m.frete_pct_min, "frete_pct_medio": m.frete_pct_medio,
        "frete_pct_max": m.frete_pct_max,
        "amostra_tamanho": m.amostra_tamanho, "metodologia": m.metodologia,
        "data_pesquisa": m.data_pesquisa.isoformat() if m.data_pesquisa else None,
    }


@router.post("/observado/gerar")
def gerar_observado(
    empresa_id: int = Query(...),
    periodo_ref: str = Query("TOTAL"),
    _=Depends(bloquear_visualizador),
    _ac=Depends(verificar_acesso_empresa),
    db: Session = Depends(get_db),
):
    """(Re)gera o benchmark observado da empresa a partir dos CT-e."""
    uc = BenchmarkObservadoUseCase(db)
    r = uc.gerar(empresa_id, periodo_ref=periodo_ref)
    return {
        "grupos_gerados": r.grupos_gerados,
        "embarques_considerados": r.embarques_considerados,
        "embarques_ignorados": r.embarques_ignorados,
        "periodo_ref": periodo_ref,
    }


@router.get("/observado")
def listar_observado(
    empresa_id: int = Query(...),
    periodo_ref: Optional[str] = Query(None),
    _=Depends(get_current_user),
    _ac=Depends(verificar_acesso_empresa),
    db: Session = Depends(get_db),
):
    uc = BenchmarkObservadoUseCase(db)
    return [_obs_dict(o) for o in uc.listar(empresa_id, periodo_ref)]


@router.get("/observado-vs-mercado")
def observado_vs_mercado(
    empresa_id: int = Query(...),
    periodo_ref: str = Query("TOTAL"),
    _=Depends(get_current_user),
    _ac=Depends(verificar_acesso_empresa),
    db: Session = Depends(get_db),
):
    return BenchmarkV2UseCase(db).observado_vs_mercado(empresa_id, periodo_ref)


@router.get("/mercado")
def listar_mercado(
    origem: Optional[str] = Query(None),
    destino: Optional[str] = Query(None),
    setor: Optional[str] = Query(None),
    _=Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    stmt = select(BenchmarkMercadoModel)
    if origem:
        stmt = stmt.where(BenchmarkMercadoModel.origem_regiao == origem.strip().upper())
    if destino:
        stmt = stmt.where(BenchmarkMercadoModel.destino_regiao == destino.strip().upper())
    if setor:
        stmt = stmt.where(BenchmarkMercadoModel.setor == setor.strip())
    stmt = stmt.order_by(BenchmarkMercadoModel.origem_regiao, BenchmarkMercadoModel.destino_regiao)
    return [_merc_dict(m) for m in db.execute(stmt).scalars().all()]


class BenchmarkMercadoIn(BaseModel):
    origem_regiao: str
    destino_regiao: str
    modal: str = ""
    tipo_operacao: str = "TODOS"
    setor: str = "TODOS"
    faixa_peso_min: float = 0.0
    faixa_peso_max: float = 0.0
    rs_kg_p10: float = Field(0.0, ge=0)
    rs_kg_p25: float = Field(0.0, ge=0)
    rs_kg_p50: float = Field(0.0, ge=0)
    rs_kg_p75: float = Field(0.0, ge=0)
    rs_kg_p90: float = Field(0.0, ge=0)
    frete_pct_min: float = Field(0.0, ge=0)
    frete_pct_medio: float = Field(0.0, ge=0)
    frete_pct_max: float = Field(0.0, ge=0)
    amostra_tamanho: Optional[int] = Field(None, ge=0)
    metodologia: str = ""
    data_pesquisa: Optional[date] = None


@router.put("/mercado")
def upsert_mercado(
    payload: BenchmarkMercadoIn,
    _=Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    """Cria ou atualiza uma linha da matriz de mercado (chave: OD + modal +
    tipo_operacao + setor + faixa de peso). Edição manual da source of truth."""
    o = payload.origem_regiao.strip().upper()
    d = payload.destino_regiao.strip().upper()
    if not o or not d:
        raise HTTPException(400, "Origem e destino são obrigatórios.")
    modal = (payload.modal or "").strip().upper()
    tipo = (payload.tipo_operacao or "TODOS").strip().upper()
    setor = (payload.setor or "TODOS").strip()

    m = db.execute(
        select(BenchmarkMercadoModel).where(
            BenchmarkMercadoModel.origem_regiao == o,
            BenchmarkMercadoModel.destino_regiao == d,
            BenchmarkMercadoModel.modal == modal,
            BenchmarkMercadoModel.tipo_operacao == tipo,
            BenchmarkMercadoModel.setor == setor,
            BenchmarkMercadoModel.faixa_peso_min == payload.faixa_peso_min,
            BenchmarkMercadoModel.faixa_peso_max == payload.faixa_peso_max,
        )
    ).scalars().first()
    if not m:
        m = BenchmarkMercadoModel(
            origem_regiao=o, destino_regiao=d, modal=modal, tipo_operacao=tipo, setor=setor,
            faixa_peso_min=payload.faixa_peso_min, faixa_peso_max=payload.faixa_peso_max,
        )
        db.add(m)
    m.rs_kg_p10 = payload.rs_kg_p10
    m.rs_kg_p25 = payload.rs_kg_p25
    m.rs_kg_p50 = payload.rs_kg_p50
    m.rs_kg_p75 = payload.rs_kg_p75
    m.rs_kg_p90 = payload.rs_kg_p90
    m.frete_pct_min = payload.frete_pct_min
    m.frete_pct_medio = payload.frete_pct_medio
    m.frete_pct_max = payload.frete_pct_max
    m.fonte = "MERCADO"
    m.amostra_tamanho = payload.amostra_tamanho
    m.metodologia = (payload.metodologia or "").strip()
    m.data_pesquisa = payload.data_pesquisa
    db.commit()
    db.refresh(m)
    return _merc_dict(m)


@router.delete("/mercado/{mid}", status_code=204)
def deletar_mercado(
    mid: int,
    _=Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    m = db.get(BenchmarkMercadoModel, mid)
    if not m:
        raise HTTPException(404, "Linha de mercado não encontrada.")
    db.delete(m)
    db.commit()


@router.get("/mercado/modelo")
def baixar_modelo_mercado(_=Depends(get_current_superuser)):
    """Baixa a planilha-modelo (.xlsx) para importação em massa da Matriz
    Mercado. Colunas: origem_regiao, destino_regiao, modal (opcional),
    tipo_operacao, setor, faixa_peso_min, faixa_peso_max, rs_kg_p10..rs_kg_p90,
    frete_pct_min/medio/max, amostra_tamanho, metodologia, data_pesquisa
    (proveniência da pesquisa)."""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz Mercado"
    colunas = [
        "origem_regiao", "destino_regiao", "modal", "tipo_operacao", "setor",
        "faixa_peso_min", "faixa_peso_max",
        "rs_kg_p10", "rs_kg_p25", "rs_kg_p50", "rs_kg_p75", "rs_kg_p90",
        "frete_pct_min", "frete_pct_medio", "frete_pct_max",
        "amostra_tamanho", "metodologia", "data_pesquisa",
    ]
    ws.append(colunas)
    azul = PatternFill("solid", fgColor="2D3561")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
    ws.append(["SUDESTE", "SUDESTE", "", "TODOS", "TODOS", 0, 0, 3.2, 3.6, 4.1, 4.7, 5.3, 1.2, 1.7, 2.3, "", "", ""])
    ws.append(["SUDESTE", "NORDESTE", "", "FTL", "TODOS", 0, 500, 4.0, 4.4, 5.0, 5.6, 6.2, 1.8, 2.6, 3.4, "", "", ""])
    ws.append(["SUDESTE", "NORDESTE", "", "FRACIONADO", "Farmacêutico", 0, 500, 9.1, 10.35, 12.4, 15.9, 19.7,
               2.5, 3.4, 4.6, 14, "Cotação direta com 6 transportadoras especializadas em carga fria", "2026-06-02"])
    for col_letra, largura in zip("ABCDEFGHIJKLMNOPQR",
                                   [14, 14, 10, 14, 16, 14, 14, 10, 10, 10, 10, 10, 12, 12, 12, 12, 40, 14]):
        ws.column_dimensions[col_letra].width = largura

    inst = wb.create_sheet("Instruções")
    for linha in [
        ["Como preencher"],
        [""],
        ["origem_regiao", "Macrorregião de origem: NORTE, NORDESTE, CENTRO_OESTE, SUDESTE ou SUL (obrigatório)."],
        ["destino_regiao", "Macrorregião de destino: mesmos valores acima (obrigatório)."],
        ["modal", "Modal de transporte (opcional), ex.: RODOVIARIO. Vazio = qualquer modal."],
        ["tipo_operacao", "TODOS, FTL, LTL ou FRACIONADO (opcional, padrão TODOS)."],
        ["setor", "Setor econômico pesquisado, ex.: Indústria, Varejo, Farmacêutico (opcional, padrão TODOS = referência genérica)."],
        ["faixa_peso_min / faixa_peso_max", "Faixa de peso em kg (opcional). 0/0 = sem faixa."],
        ["rs_kg_p10 .. rs_kg_p90", "Percentis de R$/kg de mercado para o par (obrigatório ao menos o P50)."],
        ["frete_pct_min / frete_pct_medio / frete_pct_max", "Faixa de % Frete/Mercadoria de mercado para o par (opcional)."],
        ["amostra_tamanho", "Nº de cotações/pontos de dado que sustentam a linha (opcional, recomendado)."],
        ["metodologia", "Como o dado foi coletado, texto curto (opcional, recomendado)."],
        ["data_pesquisa", "Data da coleta no formato AAAA-MM-DD (opcional, recomendado)."],
        [""],
        ["Uma linha por combinação de origem + destino + modal + tipo_operacao + setor + faixa de peso."],
        ["Linhas já cadastradas com a mesma combinação são atualizadas (upsert)."],
    ]:
        inst.append(linha)
    inst["A1"].font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 26
    inst.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_matriz_mercado.xlsx"},
    )


@router.post("/mercado/importar")
async def importar_mercado_excel(
    arquivo: UploadFile = File(...),
    _=Depends(get_current_superuser),
    db: Session = Depends(get_db),
):
    """Importa em massa linhas da Matriz Benchmark via Excel (.xlsx).
    Upsert por chave (origem+destino+modal+tipo_operacao+setor+faixa de peso),
    igual à edição manual — cada linha existente com a mesma combinação é
    atualizada, o resto é criado."""
    import datetime as dt
    import openpyxl
    from io import BytesIO
    from app.domain.entities import MacroRegiaoEnum

    conteudo = await arquivo.read()
    if not conteudo.startswith(b"PK"):
        raise HTTPException(400, "Arquivo não é um Excel (.xlsx) válido.")
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    if not {"origem_regiao", "destino_regiao"}.issubset(col):
        raise HTTPException(400, "Colunas obrigatórias ausentes: origem_regiao, destino_regiao.")

    regioes_validas = {r.value for r in MacroRegiaoEnum}

    def _num(row, campo):
        if campo not in col:
            return 0.0
        v = row[col[campo]]
        try:
            return float(v) if v not in (None, "") else 0.0
        except (TypeError, ValueError):
            return 0.0

    def _inteiro(row, campo):
        if campo not in col:
            return None
        v = row[col[campo]]
        try:
            return int(v) if v not in (None, "") else None
        except (TypeError, ValueError):
            return None

    def _data(row, campo):
        if campo not in col:
            return None
        v = row[col[campo]]
        if v in (None, ""):
            return None
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        try:
            return dt.date.fromisoformat(str(v).strip()[:10])
        except ValueError:
            return None

    importados = 0
    atualizados = 0
    ignorados = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        origem = str(row[col["origem_regiao"]] or "").strip().upper()
        destino = str(row[col["destino_regiao"]] or "").strip().upper()
        if origem not in regioes_validas or destino not in regioes_validas:
            ignorados += 1  # região inválida
            continue
        p50 = _num(row, "rs_kg_p50")
        if p50 <= 0:
            ignorados += 1  # sem mediana, referência inutilizável
            continue
        modal = str(row[col["modal"]] or "").strip().upper() if "modal" in col else ""
        tipo = (str(row[col["tipo_operacao"]] or "").strip().upper() if "tipo_operacao" in col else "") or "TODOS"
        setor = (str(row[col["setor"]] or "").strip() if "setor" in col else "") or "TODOS"
        peso_min = _num(row, "faixa_peso_min")
        peso_max = _num(row, "faixa_peso_max")

        m = db.execute(
            select(BenchmarkMercadoModel).where(
                BenchmarkMercadoModel.origem_regiao == origem,
                BenchmarkMercadoModel.destino_regiao == destino,
                BenchmarkMercadoModel.modal == modal,
                BenchmarkMercadoModel.tipo_operacao == tipo,
                BenchmarkMercadoModel.setor == setor,
                BenchmarkMercadoModel.faixa_peso_min == peso_min,
                BenchmarkMercadoModel.faixa_peso_max == peso_max,
            )
        ).scalars().first()
        if m:
            atualizados += 1
        else:
            m = BenchmarkMercadoModel(
                origem_regiao=origem, destino_regiao=destino, modal=modal, tipo_operacao=tipo, setor=setor,
                faixa_peso_min=peso_min, faixa_peso_max=peso_max,
            )
            db.add(m)
            importados += 1
        m.rs_kg_p10 = _num(row, "rs_kg_p10")
        m.rs_kg_p25 = _num(row, "rs_kg_p25")
        m.rs_kg_p50 = p50
        m.rs_kg_p75 = _num(row, "rs_kg_p75")
        m.rs_kg_p90 = _num(row, "rs_kg_p90")
        m.frete_pct_min = _num(row, "frete_pct_min")
        m.frete_pct_medio = _num(row, "frete_pct_medio")
        m.frete_pct_max = _num(row, "frete_pct_max")
        m.fonte = "MERCADO"
        m.amostra_tamanho = _inteiro(row, "amostra_tamanho")
        m.metodologia = (str(row[col["metodologia"]] or "").strip() if "metodologia" in col else "")
        m.data_pesquisa = _data(row, "data_pesquisa")

    db.commit()
    return {"importados": importados, "atualizados": atualizados, "ignorados": ignorados}


@router.get("/indicadores-regionais")
def indicadores_regionais(
    empresa_id: int = Query(...),
    _=Depends(get_current_user),
    _ac=Depends(verificar_acesso_empresa),
    db: Session = Depends(get_db),
):
    """Indicadores regionais DERIVADOS das operações reais (CT-e) — substitui o
    antigo cadastro manual de benchmark regional."""
    ind = IndicadoresRegionaisUseCase(db).calcular(empresa_id)
    return [
        {
            "regiao": i.regiao, "qtd_embarques": i.qtd_embarques,
            "frete_medio_rs_kg": i.frete_medio_rs_kg,
            "frete_faturamento_pct": i.frete_faturamento_pct,
            "peso_medio": i.peso_medio, "custo_medio": i.custo_medio,
            "lead_time_medio_dias": i.lead_time_medio_dias,
        }
        for i in ind
    ]


@router.get("/resolver")
def resolver(
    empresa_id: int = Query(...),
    origem: str = Query(...),
    destino: str = Query(...),
    modal: str = Query(""),
    valor_rs_kg: Optional[float] = Query(None),
    _=Depends(get_current_user),
    _ac=Depends(verificar_acesso_empresa),
    db: Session = Depends(get_db),
):
    """Resolve a referência (cliente→mercado). Se valor_rs_kg for informado,
    devolve também a comparação contra P50/P75."""
    uc = BenchmarkV2UseCase(db)
    if valor_rs_kg is not None:
        c = uc.comparar(empresa_id, origem, destino, valor_rs_kg, modal)
        return {
            "fonte": c.referencia.fonte, "disponivel": c.referencia.disponivel,
            "rs_kg_p50": c.referencia.rs_kg_p50, "rs_kg_p75": c.referencia.rs_kg_p75,
            "valor": c.valor, "delta_p50_pct": c.delta_p50_pct,
            "delta_p75_pct": c.delta_p75_pct, "faixa": c.faixa,
        }
    r = uc.resolver(empresa_id, origem, destino, modal)
    return {
        "fonte": r.fonte, "disponivel": r.disponivel,
        "origem_regiao": r.origem_regiao, "destino_regiao": r.destino_regiao,
        "rs_kg_p10": r.rs_kg_p10, "rs_kg_p25": r.rs_kg_p25, "rs_kg_p50": r.rs_kg_p50,
        "rs_kg_p75": r.rs_kg_p75, "rs_kg_p90": r.rs_kg_p90,
    }
