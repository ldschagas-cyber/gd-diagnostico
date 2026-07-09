"""Router do módulo Concorrência Logística — BID de Frete (V3.1).

Agrupa todos os endpoints do módulo:
  /bid                 → CRUD de BIDs
  /bid/{id}/escopo     → geração e gestão do escopo
  /bid/{id}/transp     → transportadoras convidadas
  /bid/{id}/propostas  → propostas de preço
  /bid/{id}/economia   → comparativo, motor de economia, score
  /bid/{id}/simulacoes → cenários de distribuição
  /bid/{id}/auditoria  → trilha de auditoria
  /bid/dashboard       → KPIs consolidados

Autorização (R-01): todo endpoint que opera sobre um BID existente usa a
dependency `get_bid_com_acesso`, que carrega o BID pelo `bid_id` do path e
garante que o usuário logado pertence à empresa dona dele — o `empresa_id`
efetivo para qualquer operação (nomes de transportadoras, escopo, etc.) vem
sempre de `bid.empresa_id` (autoritativo), nunca de um parâmetro de query
enviado pelo cliente.
"""
from datetime import date
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse

from app.application.use_cases.bid import BidUseCase
from app.application.use_cases.bid_economia import EconomiaUseCase, ResultadoSimulacao
from app.application.use_cases.bid_escopo import EscopoUseCase
from app.application.use_cases.benchmark_v2 import (
    BenchmarkV2UseCase, destino_regiao_de_grupo,
)
from sqlalchemy.orm import Session
from app.core.database import SessionLocal
from app.domain.entities import (
    Bid, BidAuditoria, BidProposta, BidSimulacao, BidTransportadora,
    BidStatusEnum, BidTransportadoraStatusEnum, PropostaOrigemEnum, TipoAgrupamentoEnum,
)
from app.presentation.api.dependencies import (
    bloquear_visualizador, get_bid_repo, get_bid_auditoria_repo, get_bid_com_acesso,
    get_bid_escopo_repo, get_bid_proposta_repo, get_bid_simulacao_repo,
    get_bid_transportadora_repo, get_current_user, get_db, get_transportadora_repo,
    verificar_acesso_empresa,
)
from app.presentation.schemas import (
    BidCreate, BidOut, BidStatusUpdate,
    BidEscopoOut, GerarEscopoIn,
    BidTransportadoraCreate, BidTransportadoraOut, BidTransportadoraStatusUpdate,
    BidPropostaCreate, BidPropostaOut,
    BidSimulacaoCreate, BidSimulacaoOut,
    BidAuditoriaOut,
    LinhaComparativaOut, PropostaComparativaOut, ResultadoEconomiaOut, ScoreTransportadoraOut,
    ResultadoSimulacaoOut, BidDashboardKPIOut,
)

router = APIRouter(prefix="/bid", tags=["Concorrência Logística — BID"])


# ── Helpers de injeção ────────────────────────────────────────────────────

def _bid_uc(
    bid_repo=Depends(get_bid_repo),
    audit_repo=Depends(get_bid_auditoria_repo),
) -> BidUseCase:
    return BidUseCase(bid_repo, audit_repo)


def _escopo_uc(
    db=Depends(get_db),
    escopo_repo=Depends(get_bid_escopo_repo),
) -> EscopoUseCase:
    return EscopoUseCase(db, escopo_repo)


# ════════════════════════════════════════════════════════════════
# CRUD DE BIDS
# ════════════════════════════════════════════════════════════════

@router.get("", response_model=List[BidOut])
def listar_bids(
    empresa_id: int = Query(...),
    _ac=Depends(verificar_acesso_empresa),
    uc: BidUseCase = Depends(_bid_uc),
):
    return uc.listar(empresa_id)


@router.post("", response_model=BidOut, status_code=201)
def criar_bid(
    payload: BidCreate,
    empresa_id: int = Query(...),
    user=Depends(bloquear_visualizador),
    _ac=Depends(verificar_acesso_empresa),
    uc: BidUseCase = Depends(_bid_uc),
):
    bid = Bid(empresa_id=empresa_id, **payload.model_dump())
    return uc.criar(bid, user)


@router.get("/dashboard", response_model=BidDashboardKPIOut)
def dashboard(
    empresa_id: int = Query(...),
    _ac=Depends(verificar_acesso_empresa),
    bid_repo=Depends(get_bid_repo),
    proposta_repo=Depends(get_bid_proposta_repo),
    escopo_repo=Depends(get_bid_escopo_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
):
    bids = bid_repo.list_by_empresa(empresa_id)
    total = len(bids)
    abertos = sum(1 for b in bids if b.status == BidStatusEnum.ABERTO)
    encerrados = sum(1 for b in bids if b.status == BidStatusEnum.ENCERRADO)

    eco_id = 0.0
    eco_cont = 0.0
    melhor_eco = 0.0
    transp_avaliadas = set()

    for bid in bids:
        escopos = escopo_repo.list_by_bid(bid.id)
        propostas = proposta_repo.list_by_bid(bid.id)
        bts = bt_repo.list_by_bid(bid.id)
        for bt in bts:
            transp_avaliadas.add(bt.transportadora_id)

        melhor_por_grupo = {}
        for p in propostas:
            k = p.valor_grupo
            if k not in melhor_por_grupo or p.valor_rs_kg < melhor_por_grupo[k]:
                melhor_por_grupo[k] = p.valor_rs_kg

        for e in escopos:
            if e.valor_grupo in melhor_por_grupo:
                eco = max(0.0, (e.frete_rs_kg - melhor_por_grupo[e.valor_grupo]) * e.peso_total)
                eco_id += eco
                melhor_eco = max(melhor_eco, eco)
                if bid.status == BidStatusEnum.ENCERRADO:
                    eco_cont += eco

    return BidDashboardKPIOut(
        total_bids=total,
        bids_abertos=abertos,
        bids_encerrados=encerrados,
        economia_identificada=round(eco_id, 2),
        economia_contratada=round(eco_cont, 2),
        melhor_economia=round(melhor_eco, 2),
        transportadoras_avaliadas=len(transp_avaliadas),
    )


@router.get("/{bid_id}", response_model=BidOut)
def obter_bid(bid=Depends(get_bid_com_acesso)):
    return bid


@router.put("/{bid_id}", response_model=BidOut)
def atualizar_bid(
    bid_id: int, payload: BidCreate,
    user=Depends(bloquear_visualizador),
    _bid=Depends(get_bid_com_acesso),
    uc: BidUseCase = Depends(_bid_uc),
):
    bid = Bid(**payload.model_dump())
    try:
        result = uc.atualizar(bid_id, bid, user)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not result:
        raise HTTPException(404, "BID não encontrado.")
    return result


@router.patch("/{bid_id}/status", response_model=BidOut)
def alterar_status(
    bid_id: int, payload: BidStatusUpdate,
    user=Depends(bloquear_visualizador),
    _bid=Depends(get_bid_com_acesso),
    uc: BidUseCase = Depends(_bid_uc),
):
    try:
        return uc.alterar_status(bid_id, payload.status, user)
    except ValueError as e:
        raise HTTPException(400, str(e))


@router.delete("/{bid_id}", status_code=204)
def deletar_bid(
    bid_id: int,
    user=Depends(bloquear_visualizador),
    _bid=Depends(get_bid_com_acesso),
    uc: BidUseCase = Depends(_bid_uc),
):
    try:
        if not uc.deletar(bid_id, user):
            raise HTTPException(404, "BID não encontrado.")
    except ValueError as e:
        raise HTTPException(400, str(e))


# ════════════════════════════════════════════════════════════════
# ESCOPO
# ════════════════════════════════════════════════════════════════

@router.post("/{bid_id}/escopo/gerar", response_model=List[BidEscopoOut])
def gerar_escopo(
    bid_id: int,
    payload: GerarEscopoIn,
    bid=Depends(get_bid_com_acesso),
    _=Depends(bloquear_visualizador),
    uc: EscopoUseCase = Depends(_escopo_uc),
    transp_repo=Depends(get_transportadora_repo),
):
    if not bid.periodo_analise_inicio or not bid.periodo_analise_fim:
        raise HTTPException(400, "Defina periodo_analise_inicio e periodo_analise_fim no BID antes de gerar o escopo.")

    nomes = {}
    if payload.tipo_agrupamento == TipoAgrupamentoEnum.TRANSPORTADORA:
        ts = transp_repo.list_by_empresa(bid.empresa_id, limit=1000)
        nomes = {t.id: (t.nome_fantasia or t.razao_social) for t in ts}

    return uc.gerar(
        bid_id=bid_id, empresa_id=bid.empresa_id,
        data_inicio=bid.periodo_analise_inicio, data_fim=bid.periodo_analise_fim,
        tipo_agrupamento=payload.tipo_agrupamento,
        faixas_peso=[f.model_dump() for f in payload.faixas_peso] if payload.faixas_peso else None,
        nome_transp_map=nomes,
        segmentacao_tipo=bid.segmentacao_tipo,
        segmentacao_valor=bid.segmentacao_valor,
    )


@router.get("/{bid_id}/escopo", response_model=List[BidEscopoOut])
def listar_escopo(
    bid_id: int,
    _bid=Depends(get_bid_com_acesso),
    uc: EscopoUseCase = Depends(_escopo_uc),
):
    return uc.listar(bid_id)


@router.delete("/{bid_id}/escopo/{escopo_id}", status_code=204)
def deletar_item_escopo(
    bid_id: int, escopo_id: int,
    _bid=Depends(get_bid_com_acesso),
    _=Depends(bloquear_visualizador),
    escopo_repo=Depends(get_bid_escopo_repo),
):
    if not escopo_repo.delete(escopo_id):
        raise HTTPException(404, "Item de escopo não encontrado.")


# ════════════════════════════════════════════════════════════════
# TRANSPORTADORAS CONVIDADAS
# ════════════════════════════════════════════════════════════════

@router.get("/{bid_id}/transportadoras", response_model=List[BidTransportadoraOut])
def listar_transportadoras_bid(
    bid_id: int,
    _bid=Depends(get_bid_com_acesso),
    bt_repo=Depends(get_bid_transportadora_repo),
    transp_repo=Depends(get_transportadora_repo),
):
    bts = bt_repo.list_by_bid(bid_id)
    # Resolve o nome real cadastrado de cada transportadora (evita placeholders no grid)
    nomes: dict[int, str] = {}
    for bt in bts:
        if bt.transportadora_id not in nomes:
            t = transp_repo.get(bt.transportadora_id)
            if t:
                nomes[bt.transportadora_id] = t.nome_fantasia or t.razao_social
    return [
        BidTransportadoraOut(
            id=bt.id, bid_id=bt.bid_id, transportadora_id=bt.transportadora_id,
            status=bt.status, observacoes=bt.observacoes,
            avaliacao_usuario=bt.avaliacao_usuario,
            transportadora_nome=nomes.get(bt.transportadora_id),
        )
        for bt in bts
    ]


@router.post("/{bid_id}/transportadoras", response_model=BidTransportadoraOut, status_code=201)
def convidar_transportadora(
    bid_id: int,
    payload: BidTransportadoraCreate,
    user=Depends(bloquear_visualizador),
    bid=Depends(get_bid_com_acesso),
    bt_repo=Depends(get_bid_transportadora_repo),
    transp_repo=Depends(get_transportadora_repo),
    audit_repo=Depends(get_bid_auditoria_repo),
):
    # Verifica se a transportadora pertence à empresa dona do BID
    t = transp_repo.get(payload.transportadora_id)
    if not t or t.empresa_id != bid.empresa_id:
        raise HTTPException(404, "Transportadora não encontrada nesta empresa.")
    # Evita duplicidade
    existentes = bt_repo.list_by_bid(bid_id)
    if any(e.transportadora_id == payload.transportadora_id for e in existentes):
        raise HTTPException(409, "Transportadora já está neste BID.")
    bt = BidTransportadora(bid_id=bid_id, **payload.model_dump())
    criado = bt_repo.create(bt)
    audit_repo.registrar(BidAuditoria(
        bid_id=bid_id, user_id=user.id,
        acao=f"Transportadora {t.nome_fantasia or t.razao_social} convidada",
    ))
    return criado


@router.put("/{bid_id}/transportadoras/{bt_id}", response_model=BidTransportadoraOut)
def atualizar_transportadora_bid(
    bid_id: int, bt_id: int, payload: BidTransportadoraCreate,
    _bid=Depends(get_bid_com_acesso),
    _=Depends(bloquear_visualizador),
    bt_repo=Depends(get_bid_transportadora_repo),
):
    bt = BidTransportadora(bid_id=bid_id, **payload.model_dump())
    result = bt_repo.update(bt_id, bt)
    if not result:
        raise HTTPException(404, "Participante não encontrado.")
    return result


@router.patch("/{bid_id}/transportadoras/{bt_id}/status", response_model=BidTransportadoraOut)
def status_transportadora_bid(
    bid_id: int, bt_id: int, payload: BidTransportadoraStatusUpdate,
    user=Depends(bloquear_visualizador),
    _bid=Depends(get_bid_com_acesso),
    bt_repo=Depends(get_bid_transportadora_repo),
    audit_repo=Depends(get_bid_auditoria_repo),
):
    result = bt_repo.update_status(bt_id, payload.status)
    if not result:
        raise HTTPException(404, "Participante não encontrado.")
    audit_repo.registrar(BidAuditoria(
        bid_id=bid_id, user_id=user.id,
        acao=f"Transportadora (id {result.transportadora_id}) → {payload.status.value}",
    ))
    return result


@router.delete("/{bid_id}/transportadoras/{bt_id}", status_code=204)
def remover_transportadora_bid(
    bid_id: int, bt_id: int,
    _bid=Depends(get_bid_com_acesso),
    _=Depends(bloquear_visualizador),
    bt_repo=Depends(get_bid_transportadora_repo),
):
    if not bt_repo.delete(bt_id):
        raise HTTPException(404, "Participante não encontrado.")


# ════════════════════════════════════════════════════════════════
# PROPOSTAS
# ════════════════════════════════════════════════════════════════

@router.get("/{bid_id}/propostas", response_model=List[BidPropostaOut])
def listar_propostas(
    bid_id: int,
    _bid=Depends(get_bid_com_acesso),
    repo=Depends(get_bid_proposta_repo),
):
    return repo.list_by_bid(bid_id)


@router.post("/{bid_id}/propostas", response_model=BidPropostaOut, status_code=201)
def incluir_proposta(
    bid_id: int, payload: BidPropostaCreate,
    user=Depends(bloquear_visualizador),
    _bid=Depends(get_bid_com_acesso),
    repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    audit_repo=Depends(get_bid_auditoria_repo),
):
    # Valida que bid_transportadora pertence ao bid
    bt = bt_repo.get(payload.bid_transportadora_id)
    if not bt or bt.bid_id != bid_id:
        raise HTTPException(404, "Participante não pertence a este BID.")
    proposta = BidProposta(bid_id=bid_id, **payload.model_dump())
    criado = repo.create(proposta)
    # Atualiza status do participante
    bt_repo.update_status(bt.id, BidTransportadoraStatusEnum.PROPOSTA_RECEBIDA)
    audit_repo.registrar(BidAuditoria(
        bid_id=bid_id, user_id=user.id,
        acao=f"Proposta incluída — {payload.valor_grupo} @ R${payload.valor_rs_kg}/kg",
    ))
    return criado


@router.get("/{bid_id}/propostas/modelo")
def baixar_modelo_propostas(
    bid_id: int,
    _=Depends(get_current_user),
    bid=Depends(get_bid_com_acesso),
    escopo_repo=Depends(get_bid_escopo_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    transp_repo=Depends(get_transportadora_repo),
):
    """Gera a planilha-modelo de propostas para enviar às transportadoras.

    Vem pré-preenchida com os grupos do escopo do BID (um por linha) cruzados
    com cada transportadora convidada, deixando o preço em branco para a
    transportadora completar. O arquivo é diretamente compatível com a
    importação de propostas (mesmas colunas obrigatórias).
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    escopos = escopo_repo.list_by_bid(bid_id)
    bts = bt_repo.list_by_bid(bid_id)

    # nomes reais das transportadoras convidadas
    convidadas = []
    for bt in bts:
        t = transp_repo.get(bt.transportadora_id)
        nome = (t.nome_fantasia or t.razao_social) if t else f"Transportadora {bt.transportadora_id}"
        convidadas.append((bt.transportadora_id, nome))

    wb = Workbook()
    ws = wb.active
    ws.title = "Propostas"

    colunas = [
        "transportadora_id", "transportadora", "tipo_agrupamento", "valor_grupo",
        "valor_rs_kg", "valor_minimo", "prazo_dias", "cobertura_pct", "observacoes",
    ]
    # cabeçalho das colunas que a transportadora preenche fica destacado
    preencher = {"valor_rs_kg", "valor_minimo", "prazo_dias", "cobertura_pct", "observacoes"}
    fill_fixo = PatternFill("solid", fgColor="2D3561")
    fill_preencher = PatternFill("solid", fgColor="C9A84C")
    for c, nome in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=c, value=nome)
        cel.fill = fill_preencher if nome in preencher else fill_fixo
        cel.font = Font(bold=True, color="FFFFFF")
        cel.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cel.column_letter].width = max(14, len(nome) + 2)

    # linhas pré-preenchidas: cada transportadora × cada grupo do escopo
    linha = 2
    pares_transp = convidadas or [("", "")]  # se nenhuma convidada, deixa em branco
    for tid, tnome in pares_transp:
        for e in escopos:
            ws.cell(row=linha, column=1, value=tid)
            ws.cell(row=linha, column=2, value=tnome)
            ws.cell(row=linha, column=3, value=e.tipo_agrupamento)
            ws.cell(row=linha, column=4, value=e.valor_grupo)
            ws.cell(row=linha, column=8, value=100)  # cobertura_pct padrão
            linha += 1

    inst = wb.create_sheet("Instruções")
    txt = [
        ("Modelo de Propostas — " + bid.nome, True),
        ("", False),
        ("Para a transportadora:", True),
        ("- Preencha apenas as colunas em destaque dourado:", False),
        ("  valor_rs_kg (R$/kg), valor_minimo, prazo_dias, cobertura_pct, observacoes.", False),
        ("- NÃO altere as colunas transportadora_id, tipo_agrupamento e valor_grupo.", False),
        ("- Cada linha corresponde a um trecho/grupo a ser cotado.", False),
        ("- Valores em R$ aceitam vírgula decimal (ex.: 1,85).", False),
        ("", False),
        ("Para importar de volta no sistema:", True),
        ("- Use a aba “Importar Excel” dentro do BID.", False),
        ("- O arquivo já está no formato correto de importação.", False),
    ]
    for i, (t, b) in enumerate(txt, start=1):
        cel = inst.cell(row=i, column=1, value=t)
        if b:
            cel.font = Font(bold=True)
    inst.column_dimensions["A"].width = 84

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    nome_arq = f"modelo_propostas_bid_{bid_id}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{nome_arq}"'},
    )


@router.post("/{bid_id}/propostas/importar", response_model=dict)
async def importar_propostas_excel(
    bid_id: int,
    arquivo: UploadFile = File(...),
    user=Depends(bloquear_visualizador),
    _bid=Depends(get_bid_com_acesso),
    repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    audit_repo=Depends(get_bid_auditoria_repo),
):
    """Importação de propostas via Excel.

    O arquivo deve ter as colunas (em qualquer ordem):
    transportadora_id | tipo_agrupamento | valor_grupo | valor_rs_kg | valor_minimo | prazo_dias | cobertura_pct | observacoes
    """
    import openpyxl

    conteudo = await arquivo.read()
    if not conteudo.startswith(b"PK"):
        raise HTTPException(400, "Arquivo não é um Excel (.xlsx) válido.")

    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    obrigatorias = {"transportadora_id", "tipo_agrupamento", "valor_grupo", "valor_rs_kg"}
    faltando = obrigatorias - set(col.keys())
    if faltando:
        raise HTTPException(400, f"Colunas obrigatórias ausentes: {faltando}")

    bts_do_bid = {bt.transportadora_id: bt for bt in bt_repo.list_by_bid(bid_id)}
    propostas = []
    ignoradas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        try:
            tid = int(row[col["transportadora_id"]])
        except (TypeError, ValueError):
            ignoradas += 1
            continue
        bt = bts_do_bid.get(tid)
        if not bt:
            ignoradas += 1  # transportadora não está convidada no BID
            continue
        # Sem preço não há proposta a comparar
        rs_kg = float(row[col.get("valor_rs_kg", -1)] or 0)
        if rs_kg <= 0:
            ignoradas += 1
            continue

        tipo_str = str(row[col["tipo_agrupamento"]]).upper().strip()
        try:
            tipo = TipoAgrupamentoEnum(tipo_str)
        except ValueError:
            tipo = TipoAgrupamentoEnum.REGIAO

        propostas.append(BidProposta(
            bid_id=bid_id,
            bid_transportadora_id=bt.id,
            tipo_agrupamento=tipo,
            valor_grupo=str(row[col["valor_grupo"]] or ""),
            valor_rs_kg=rs_kg,
            valor_minimo=float(row[col.get("valor_minimo", -1)] or 0) if "valor_minimo" in col else 0.0,
            prazo_dias=int(row[col.get("prazo_dias", -1)] or 0) if "prazo_dias" in col else 0,
            cobertura_pct=float(row[col.get("cobertura_pct", -1)] or 100) if "cobertura_pct" in col else 100.0,
            observacoes=str(row[col.get("observacoes", -1)] or "") if "observacoes" in col else "",
            origem=PropostaOrigemEnum.EXCEL,
        ))

    importados = repo.bulk_create(propostas)
    audit_repo.registrar(BidAuditoria(
        bid_id=bid_id, user_id=user.id,
        acao=f"{importados} propostas importadas via Excel ({ignoradas} ignoradas)",
    ))
    return {"importadas": importados, "ignoradas": ignoradas}


@router.delete("/{bid_id}/propostas/{pid}", status_code=204)
def deletar_proposta(
    bid_id: int, pid: int,
    _bid=Depends(get_bid_com_acesso),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_bid_proposta_repo),
):
    if not repo.delete(pid):
        raise HTTPException(404, "Proposta não encontrada.")


# ════════════════════════════════════════════════════════════════
# COMPARATIVO, ECONOMIA E SCORE
# ════════════════════════════════════════════════════════════════

@router.get("/{bid_id}/comparativo", response_model=List[LinhaComparativaOut])
def comparativo(
    bid_id: int,
    bid=Depends(get_bid_com_acesso),
    escopo_repo=Depends(get_bid_escopo_repo),
    proposta_repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    simulacao_repo=Depends(get_bid_simulacao_repo),
    transp_repo=Depends(get_transportadora_repo),
    db: Session = Depends(get_db),
):
    ts = transp_repo.list_by_empresa(bid.empresa_id, limit=1000)
    nomes = {t.id: (t.nome_fantasia or t.razao_social) for t in ts}
    uc = EconomiaUseCase(escopo_repo, proposta_repo, bt_repo, simulacao_repo, nomes)
    linhas = uc.comparativo(bid_id)

    # Enriquecimento Benchmark V2 — referência de mercado (P50/P75) por destino.
    bv2 = BenchmarkV2UseCase(db)
    escopos = escopo_repo.list_by_bid(bid_id)
    tipo_por_grupo = {e.valor_grupo: e.tipo_agrupamento for e in escopos}
    ref_cache: dict[str, object] = {}

    resultado: List[LinhaComparativaOut] = []
    for ln in linhas:
        tipo = tipo_por_grupo.get(ln.valor_grupo, "")
        destino = destino_regiao_de_grupo(tipo, ln.valor_grupo)
        ref = None
        if destino:
            ref = ref_cache.get(destino) or bv2.referencia_destino(destino)
            ref_cache[destino] = ref

        props_out = []
        for p in ln.propostas:
            faixa = delta = None
            if ref is not None and ref.disponivel:
                faixa = BenchmarkV2UseCase.faixa_de(p["rs_kg"], ref.rs_kg_p50, ref.rs_kg_p75)
                delta = round((p["rs_kg"] - ref.rs_kg_p50) / ref.rs_kg_p50 * 100, 2) if ref.rs_kg_p50 else None
            props_out.append(PropostaComparativaOut(
                bid_transportadora_id=p["bid_transportadora_id"],
                transportadora=p["transportadora"], rs_kg=p["rs_kg"],
                valor_minimo=p["valor_minimo"], prazo_dias=p["prazo_dias"],
                cobertura_pct=p["cobertura_pct"], economia_rs=p["economia_rs"],
                economia_pct=p["economia_pct"],
                faixa_mercado=faixa, delta_mercado_p50_pct=delta,
            ))

        resultado.append(LinhaComparativaOut(
            valor_grupo=ln.valor_grupo, peso_total=ln.peso_total,
            frete_atual_rs_kg=ln.frete_atual_rs_kg, custo_atual_total=ln.custo_atual_total,
            propostas=props_out,
            melhor_proposta_rs_kg=ln.melhor_proposta_rs_kg, melhor_economia=ln.melhor_economia,
            mercado_p50=(ref.rs_kg_p50 if ref and ref.disponivel else None),
            mercado_p75=(ref.rs_kg_p75 if ref and ref.disponivel else None),
            mercado_fonte=(ref.fonte if ref else None),
        ))
    return resultado


@router.get("/{bid_id}/economia", response_model=ResultadoEconomiaOut)
def motor_economia(
    bid_id: int,
    bid=Depends(get_bid_com_acesso),
    escopo_repo=Depends(get_bid_escopo_repo),
    proposta_repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    simulacao_repo=Depends(get_bid_simulacao_repo),
    transp_repo=Depends(get_transportadora_repo),
):
    ts = transp_repo.list_by_empresa(bid.empresa_id, limit=1000)
    nomes = {t.id: (t.nome_fantasia or t.razao_social) for t in ts}
    uc = EconomiaUseCase(escopo_repo, proposta_repo, bt_repo, simulacao_repo, nomes)
    return uc.economia(bid_id, bid.periodo_analise_inicio, bid.periodo_analise_fim)


@router.get("/{bid_id}/score", response_model=List[ScoreTransportadoraOut])
def score_transportadoras(
    bid_id: int,
    bid=Depends(get_bid_com_acesso),
    escopo_repo=Depends(get_bid_escopo_repo),
    proposta_repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    simulacao_repo=Depends(get_bid_simulacao_repo),
    transp_repo=Depends(get_transportadora_repo),
):
    ts = transp_repo.list_by_empresa(bid.empresa_id, limit=1000)
    nomes = {t.id: (t.nome_fantasia or t.razao_social) for t in ts}
    uc = EconomiaUseCase(escopo_repo, proposta_repo, bt_repo, simulacao_repo, nomes)
    return uc.scores(bid_id)


# ════════════════════════════════════════════════════════════════
# SIMULAÇÃO DE DISTRIBUIÇÃO
# ════════════════════════════════════════════════════════════════

@router.get("/{bid_id}/simulacoes", response_model=List[BidSimulacaoOut])
def listar_simulacoes(
    bid_id: int,
    _bid=Depends(get_bid_com_acesso),
    simulacao_repo=Depends(get_bid_simulacao_repo),
):
    return simulacao_repo.list_by_bid(bid_id)


@router.post("/{bid_id}/simulacoes/calcular", response_model=ResultadoSimulacaoOut)
def calcular_simulacao(
    bid_id: int,
    payload: BidSimulacaoCreate,
    bid=Depends(get_bid_com_acesso),
    escopo_repo=Depends(get_bid_escopo_repo),
    proposta_repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    simulacao_repo=Depends(get_bid_simulacao_repo),
    transp_repo=Depends(get_transportadora_repo),
):
    ts = transp_repo.list_by_empresa(bid.empresa_id, limit=1000)
    nomes = {t.id: (t.nome_fantasia or t.razao_social) for t in ts}
    uc = EconomiaUseCase(escopo_repo, proposta_repo, bt_repo, simulacao_repo, nomes)
    itens = [i.model_dump() for i in payload.itens]
    return uc.calcular_simulacao(bid_id, itens)


@router.post("/{bid_id}/simulacoes", response_model=BidSimulacaoOut, status_code=201)
def salvar_simulacao(
    bid_id: int,
    payload: BidSimulacaoCreate,
    user=Depends(bloquear_visualizador),
    bid=Depends(get_bid_com_acesso),
    escopo_repo=Depends(get_bid_escopo_repo),
    proposta_repo=Depends(get_bid_proposta_repo),
    bt_repo=Depends(get_bid_transportadora_repo),
    simulacao_repo=Depends(get_bid_simulacao_repo),
    transp_repo=Depends(get_transportadora_repo),
    audit_repo=Depends(get_bid_auditoria_repo),
):
    ts = transp_repo.list_by_empresa(bid.empresa_id, limit=1000)
    nomes = {t.id: (t.nome_fantasia or t.razao_social) for t in ts}
    uc = EconomiaUseCase(escopo_repo, proposta_repo, bt_repo, simulacao_repo, nomes)
    itens = [i.model_dump() for i in payload.itens]
    resultado = uc.calcular_simulacao(bid_id, itens)
    sim = BidSimulacao(
        bid_id=bid_id, nome=payload.nome, descricao=payload.descricao,
        itens=itens, custo_atual=resultado.custo_atual,
        custo_proposto=resultado.custo_proposto,
        economia_total=resultado.economia_total,
        reducao_pct=resultado.reducao_pct,
    )
    criado = uc.salvar_simulacao(sim)
    audit_repo.registrar(BidAuditoria(
        bid_id=bid_id, user_id=user.id,
        acao=f"Simulação '{payload.nome}' salva — economia R${resultado.economia_total:,.2f}",
    ))
    return criado


@router.delete("/{bid_id}/simulacoes/{sim_id}", status_code=204)
def deletar_simulacao(
    bid_id: int, sim_id: int,
    _bid=Depends(get_bid_com_acesso),
    _=Depends(bloquear_visualizador),
    simulacao_repo=Depends(get_bid_simulacao_repo),
):
    if not simulacao_repo.delete(sim_id):
        raise HTTPException(404, "Simulação não encontrada.")


# ════════════════════════════════════════════════════════════════
# AUDITORIA
# ════════════════════════════════════════════════════════════════

@router.get("/{bid_id}/auditoria", response_model=List[BidAuditoriaOut])
def trilha_auditoria(
    bid_id: int,
    _bid=Depends(get_bid_com_acesso),
    audit_repo=Depends(get_bid_auditoria_repo),
):
    return audit_repo.list_by_bid(bid_id)
