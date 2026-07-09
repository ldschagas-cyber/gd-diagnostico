"""Router de regiões (RF005) e cidades (RF006), com importação CSV."""
import csv
import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from app.domain.entities import Cidade, MacroRegiaoEnum, Regiao
from app.presentation.api.dependencies import (
    bloquear_visualizador,
    get_cidade_repo,
    get_current_user,
    get_regiao_repo,
)
from app.presentation.schemas import (
    CidadeCreate,
    CidadeOut,
    RegiaoCreate,
    RegiaoOut,
)

router = APIRouter(tags=["Regiões e Cidades"])


# ---------------- Regiões (RF005) ----------------
@router.get("/regioes", response_model=list[RegiaoOut])
def listar_regioes(_=Depends(get_current_user), repo=Depends(get_regiao_repo)):
    return repo.list(limit=1000)


@router.post("/regioes", response_model=RegiaoOut, status_code=201)
def criar_regiao(payload: RegiaoCreate, _=Depends(bloquear_visualizador), repo=Depends(get_regiao_repo)):
    return repo.create(Regiao(**payload.model_dump()))


@router.put("/regioes/{rid}", response_model=RegiaoOut)
def atualizar_regiao(
    rid: int, payload: RegiaoCreate, _=Depends(bloquear_visualizador), repo=Depends(get_regiao_repo),
):
    if not repo.get(rid):
        raise HTTPException(404, "Região não encontrada")
    return repo.update(rid, Regiao(id=rid, **payload.model_dump()))


@router.delete("/regioes/{rid}", status_code=204)
def remover_regiao(rid: int, _=Depends(bloquear_visualizador), repo=Depends(get_regiao_repo)):
    if not repo.delete(rid):
        raise HTTPException(404, "Região não encontrada")


@router.post("/regioes/importar-csv", response_model=dict)
async def importar_regioes_csv(
    arquivo: UploadFile = File(...),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_regiao_repo),
):
    """Importa regiões via CSV com colunas: nome, descricao (RF005)."""
    conteudo = (await arquivo.read()).decode("utf-8-sig")
    leitor = csv.DictReader(io.StringIO(conteudo))
    criadas = 0
    for linha in leitor:
        nome = (linha.get("nome") or linha.get("Nome") or "").strip()
        if not nome:
            continue
        repo.create(Regiao(nome=nome, descricao=(linha.get("descricao") or "").strip()))
        criadas += 1
    return {"criadas": criadas}


# ---------------- Cidades (RF006) ----------------
@router.get("/cidades", response_model=list[CidadeOut])
def listar_cidades(_=Depends(get_current_user), repo=Depends(get_cidade_repo)):
    return repo.list()


@router.post("/cidades", response_model=CidadeOut, status_code=201)
def criar_cidade(payload: CidadeCreate, _=Depends(bloquear_visualizador), repo=Depends(get_cidade_repo)):
    return repo.create(Cidade(**payload.model_dump()))


@router.put("/cidades/{cid}", response_model=CidadeOut)
def atualizar_cidade(
    cid: int, payload: CidadeCreate, _=Depends(bloquear_visualizador), repo=Depends(get_cidade_repo),
):
    if not repo.get(cid):
        raise HTTPException(404, "Cidade não encontrada")
    return repo.update(cid, Cidade(id=cid, **payload.model_dump()))


@router.delete("/cidades/{cid}", status_code=204)
def remover_cidade(cid: int, _=Depends(bloquear_visualizador), repo=Depends(get_cidade_repo)):
    if not repo.delete(cid):
        raise HTTPException(404, "Cidade não encontrada")


# ---------------- Cidades — modelo e importação Excel (RF006) ----------------

@router.get("/cidades/modelo")
def baixar_modelo_cidades(_=Depends(get_current_user)):
    """Baixa a planilha-modelo (.xlsx) para importação de cidades.

    Colunas: nome, uf, regiao_logistica (opcional). A macrorregião é derivada
    automaticamente da UF; a região logística é casada pelo nome (se existir)."""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cidades"
    headers = ["nome", "uf", "regiao_logistica"]
    ws.append(headers)
    azul = PatternFill("solid", fgColor="2D3561")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = azul
    # Exemplos
    ws.append(["São Paulo", "SP", "Interior SP"])
    ws.append(["Campinas", "SP", "Interior SP"])
    ws.append(["Recife", "PE", ""])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 24

    inst = wb.create_sheet("Instruções")
    for i, linha in enumerate([
        ["Como preencher"],
        [""],
        ["nome", "Nome da cidade (obrigatório)."],
        ["uf", "Sigla da UF com 2 letras, ex.: SP, PE (obrigatório)."],
        ["regiao_logistica", "Nome de uma Região Logística já cadastrada (opcional)."],
        [""],
        ["A macrorregião (Sudeste, Nordeste...) é preenchida automaticamente pela UF."],
        ["Cidades já cadastradas (mesmo nome + UF) são ignoradas."],
    ], start=1):
        inst.append(linha)
    inst["A1"].font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 22
    inst.column_dimensions["B"].width = 70

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_cidades.xlsx"},
    )


@router.post("/cidades/importar", response_model=dict)
async def importar_cidades_excel(
    arquivo: UploadFile = File(...),
    _=Depends(bloquear_visualizador),
    repo=Depends(get_cidade_repo),
    regiao_repo=Depends(get_regiao_repo),
):
    """Importa cidades via Excel (.xlsx). Colunas: nome, uf, regiao_logistica.
    Deriva a macrorregião pela UF e ignora cidades já cadastradas (nome+UF)."""
    import openpyxl
    from io import BytesIO
    from app.infrastructure.parsers.macro_regiao import macro_por_uf

    conteudo = await arquivo.read()
    if not conteudo.startswith(b"PK"):
        raise HTTPException(400, "Arquivo não é um Excel (.xlsx) válido.")
    wb = openpyxl.load_workbook(BytesIO(conteudo), data_only=True)
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    if "nome" not in col or "uf" not in col:
        raise HTTPException(400, "Colunas obrigatórias ausentes: nome, uf.")

    # Mapa nome(lower) -> regiao_id das regiões logísticas cadastradas
    regioes = {r.nome.strip().lower(): r.id for r in regiao_repo.list()}

    importadas = 0
    ignoradas = 0
    vistos: set[tuple[str, str]] = set()  # dedup dentro do próprio arquivo
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        nome = str(row[col["nome"]] or "").strip()
        uf = str(row[col["uf"]] or "").strip().upper()
        macro = macro_por_uf(uf)
        if not nome or macro is None:
            ignoradas += 1  # nome vazio ou UF inexistente
            continue
        chave = (nome.lower(), uf)
        if chave in vistos or repo.get_by_nome_uf(nome, uf):
            ignoradas += 1  # já cadastrada ou repetida na planilha
            continue
        vistos.add(chave)
        regiao_id = None
        if "regiao_logistica" in col:
            rnome = str(row[col["regiao_logistica"]] or "").strip().lower()
            if rnome:
                regiao_id = regioes.get(rnome)
        repo.create(Cidade(nome=nome, uf=uf, regiao_id=regiao_id, macro_regiao=macro))
        importadas += 1

    return {"importadas": importadas, "ignoradas": ignoradas}
